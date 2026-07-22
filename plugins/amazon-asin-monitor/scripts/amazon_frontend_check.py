import html
import hashlib
import json
import re
import shutil
import subprocess
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from ctypes import Structure, byref, c_ulong, c_ulonglong, sizeof, windll
from dataclasses import dataclass, replace
from datetime import datetime
from os import cpu_count, environ
from pathlib import Path
from threading import Lock
from time import perf_counter
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, ImageDraw, ImageFont

try:
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
    from playwright.sync_api import sync_playwright
except Exception:  # pragma: no cover - dependency guard for scheduled task logs
    PlaywrightTimeoutError = Exception
    sync_playwright = None


LEGACY_ROOT = Path(r"D:\Codex")
ROOT = LEGACY_ROOT / "各项目链接检查"
DESKTOP_CANDIDATES = [Path.home() / "Desktop", Path(r"D:\桌面")]
DESKTOP_INPUT_SUBDIRS = ["ASIN检查基础信息"]
INPUT_PATTERN = "*-ASIN检查基础信息.xlsx"
DATE = datetime.now().strftime("%Y-%m-%d")
DATE_DOT = datetime.now().strftime("%Y.%m.%d")
STAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
RANK_RED_THRESHOLD = 0.10
DEFAULT_POSTAL_CODE = "10043"
POSTAL_CODES = [code.strip() for code in environ.get("AMAZON_POSTAL_CODES", "10043").split(",") if code.strip()]
PROJECT_FILTER = environ.get("ASIN_PROJECT_FILTER", "").strip()
PROJECT_START = environ.get("ASIN_PROJECT_START", "").strip()
ASIN_LIMIT = int(environ.get("ASIN_LIMIT", "0") or "0")
ASIN_PARENT_FILTER = environ.get("ASIN_PARENT_FILTER", "").strip()
ASIN_CHILD_FILTER = {asin.strip().upper() for asin in environ.get("ASIN_CHILD_FILTER", "").split(",") if asin.strip()}
MAX_CONSECUTIVE_BLOCKS = int(environ.get("ASIN_MAX_CONSECUTIVE_BLOCKS", "5"))
SKIP_PREFLIGHT = environ.get("ASIN_SKIP_PREFLIGHT", "").strip().lower() in {"1", "true", "yes", "y"}
ASIN_BATCH_SIZE = max(1, int(environ.get("ASIN_BATCH_SIZE", "25") or "25"))
ASIN_BATCH_WORKERS = max(0, int(environ.get("ASIN_BATCH_WORKERS", "0") or "0"))
ASIN_BATCH_RESUME = environ.get("ASIN_BATCH_RESUME", "1").strip().lower() not in {"0", "false", "no", "n"}
MAX_BATCH_WORKERS = 3
CURL_USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
CURL_ACCEPT = "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"


def console_log(scope: str, message: str, level: str = "INFO") -> None:
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] [{level}] [{scope}] {message}"
    try:
        print(line, flush=True)
    except (OSError, ValueError):
        # A detached or expired Codex console must never stop the crawler.
        pass
    try:
        with (LEGACY_ROOT / "amazon_monitor.log").open("a", encoding="utf-8") as log_file:
            log_file.write(line + "\n")
    except (OSError, ValueError):
        pass

REPORT_HEADERS = [
    "日期/检查内容", "父ASIN", "子SKU", "子ASIN", "是否异常", "购物车", "文案", "A+是否显示",
    "评分星级", "评论数", "评论新增数", "星级占比截图", "是否来新的差评", "跟卖情况", "类目节点",
    "大类排名", "小类排名", "List Price", "Typical Price", "是否有划线", "当前/Buy Box价格",
    "划线百分比", "Prime专享折扣", "Coupon", "买赠/多买折扣", "价格截图", "备注", "本周异常是否处理完毕",
]

EXCEPTION_HEADERS = [
    "日期", "父ASIN", "子ASIN", "健康状态", "最高优先级", "问题模块", "是否需处理", "问题摘要",
    "关键变化", "证据/截图", "建议动作", "责任跟进项", "处理状态", "备注",
]


@dataclass
class ProjectPaths:
    name: str
    root: Path
    input_dir: Path
    output_dir: Path
    cache_dir: Path
    input_file: Path
    total_book: Path
    exception_book: Path
    snapshot_file: Path
    screenshot_dir: Path
    script_cache_dir: Path


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = html.unescape(str(value))
    text = re.sub(r"<script\b.*?</script>|<style\b.*?</style>", " ", text, flags=re.I | re.S)
    text = re.sub(r"<[^>]+>", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def concise_offer_text(value: Any, limit: int = 280) -> str:
    text = clean_text(value)
    if len(text) > limit or any(marker in text for marker in ["offerListingId", "marketplaceOffers", '\\":{']):
        return ""
    return text


def normalize_unavailable_offer_fields(current: dict[str, Any]) -> dict[str, Any]:
    buybox = str(current.get("buybox", ""))
    if "不可售" not in buybox and "无购物车" not in buybox:
        return current
    for field in ["list_price", "typical_price", "current_price", "discount", "prime", "coupon", "multi_buy"]:
        current[field] = ""
    current["has_strike"] = "否"
    current["price_context"] = ""
    return current


def safe_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value or "unknown").strip("_") or "unknown"


ITEM_KEY_ALIASES = {
    "父ASIN": ("父ASIN", "鐖禔SIN"),
    "子ASIN": ("子ASIN", "瀛怉SIN"),
    "子SKU": ("子SKU", "瀛怱KU"),
    "子ASIN网址": ("子ASIN网址", "瀛怉SIN缃戝潃"),
    "父ASIN网址": ("父ASIN网址", "鐖禔SIN缃戝潃"),
    "备注": ("备注", "澶囨敞"),
}


def item_value(item: dict[str, Any], key: str, default: str = "") -> str:
    for alias in ITEM_KEY_ALIASES.get(key, (key,)):
        value = item.get(alias)
        if value not in (None, ""):
            return str(value)
    return default


def with_item_aliases(item: dict[str, str]) -> dict[str, str]:
    enriched = dict(item)
    for key, aliases in ITEM_KEY_ALIASES.items():
        value = item_value(enriched, key, "")
        for alias in aliases:
            enriched.setdefault(alias, value)
    return enriched


def parse_int(value: Any) -> int:
    match = re.search(r"\d[\d,]*", str(value or ""))
    return int(match.group(0).replace(",", "")) if match else 0


def parse_money_value(value: Any) -> float:
    match = re.search(r"\d[\d,]*(?:\.\d{2})?", str(value or ""))
    return float(match.group(0).replace(",", "")) if match else 0.0


def asin_from_url(url: str) -> str:
    match = re.search(r"/(?:dp|gp/product)/([A-Z0-9]{10})", url or "", flags=re.I)
    return match.group(1).upper() if match else ""


def money(value: str) -> str:
    match = re.search(r"\$\s*[\d,]+(?:\.\d{2})?", clean_text(value))
    return match.group(0).replace(" ", "") if match else clean_text(value)


def first_money(value: str) -> str:
    match = re.search(r"\$\s*[\d,]+(?:\.\d{2})?", clean_text(value))
    return match.group(0).replace(" ", "") if match else ""


def desktop_input_files() -> list[Path]:
    ROOT.mkdir(parents=True, exist_ok=True)
    files = []
    seen = set()
    for desktop in DESKTOP_CANDIDATES:
        if desktop.exists():
            search_roots = [desktop]
            for name in DESKTOP_INPUT_SUBDIRS:
                nested = desktop / name
                if nested.is_dir():
                    search_roots.append(nested)
            for search_root in search_roots:
                for source in search_root.glob(INPUT_PATTERN):
                    resolved = source.resolve()
                    if resolved in seen:
                        continue
                    seen.add(resolved)
                    files.append(source)
    return sorted(files, key=lambda p: p.stat().st_mtime)


def ensure_legacy_project_entry(project_root: Path) -> None:
    LEGACY_ROOT.mkdir(parents=True, exist_ok=True)
    link_path = LEGACY_ROOT / project_root.name
    if link_path.resolve() == project_root.resolve():
        return
    if link_path.exists():
        try:
            if link_path.resolve() == project_root.resolve():
                subprocess.run(
                    ["attrib", "+h", "+s", "/l", str(link_path)],
                    check=False,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
                return
        except Exception:
            return
        return
    try:
        subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-Command",
                "param([string]$link,[string]$target) New-Item -ItemType Junction -Path $link -Target $target | Out-Null",
                str(link_path),
                str(project_root),
            ],
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        subprocess.run(
            ["attrib", "+h", "+s", "/l", str(link_path)],
            check=False,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
    except Exception:
        pass


def ensure_project_from_desktop_file(source: Path) -> ProjectPaths:
    project_name = source.name.removesuffix("-ASIN检查基础信息.xlsx").strip()
    if not project_name:
        raise ValueError(f"无法从文件名提取项目名称: {source}")
    project_root = ROOT / project_name
    input_dir = project_root / "1_输入需求信息"
    output_dir = project_root / "2_输出信息"
    cache_dir = project_root / "3_系统运行缓存"
    for path in (input_dir, output_dir, cache_dir):
        path.mkdir(parents=True, exist_ok=True)
    target = input_dir / source.name
    if source.resolve() != target.resolve():
        if target.exists():
            archived = input_dir / f"{target.stem}_{STAMP}{target.suffix}"
            target.rename(archived)
        shutil.move(str(source), str(target))
    script_cache_dir = cache_dir / "scripts"
    script_cache_dir.mkdir(parents=True, exist_ok=True)
    try:
        shutil.copy2(Path(__file__), script_cache_dir / Path(__file__).name)
    except Exception:
        pass
    ensure_legacy_project_entry(project_root)
    return ProjectPaths(
        name=project_name,
        root=project_root,
        input_dir=input_dir,
        output_dir=output_dir,
        cache_dir=cache_dir,
        input_file=target,
        total_book=output_dir / "ASIN检查总表.xlsx",
        exception_book=output_dir / "ASIN异常汇总表.xlsx",
        snapshot_file=cache_dir / "latest_snapshot.json",
        screenshot_dir=cache_dir / "screenshots" / DATE,
        script_cache_dir=script_cache_dir,
    )


def project_paths_from_existing(project_root: Path) -> ProjectPaths | None:
    input_dir = project_root / "1_输入需求信息"
    output_dir = project_root / "2_输出信息"
    cache_dir = project_root / "3_系统运行缓存"
    if not (input_dir.is_dir() and output_dir.is_dir() and cache_dir.is_dir()):
        return None
    inputs = sorted(input_dir.glob(INPUT_PATTERN), key=lambda p: p.stat().st_mtime, reverse=True)
    if not inputs:
        inputs = sorted(input_dir.glob("*ASIN检查基础信息.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not inputs:
        return None
    script_cache_dir = cache_dir / "scripts"
    script_cache_dir.mkdir(parents=True, exist_ok=True)
    ensure_legacy_project_entry(project_root)
    return ProjectPaths(
        name=project_root.name,
        root=project_root,
        input_dir=input_dir,
        output_dir=output_dir,
        cache_dir=cache_dir,
        input_file=inputs[0],
        total_book=output_dir / "ASIN检查总表.xlsx",
        exception_book=output_dir / "ASIN异常汇总表.xlsx",
        snapshot_file=cache_dir / "latest_snapshot.json",
        screenshot_dir=cache_dir / "screenshots" / DATE,
        script_cache_dir=script_cache_dir,
    )


def discover_projects() -> list[ProjectPaths]:
    projects: dict[Path, ProjectPaths] = {}
    for source in desktop_input_files():
        paths = ensure_project_from_desktop_file(source)
        projects[paths.root] = paths
    for child in ROOT.iterdir():
        if not child.is_dir():
            continue
        paths = project_paths_from_existing(child)
        if paths:
            projects.setdefault(paths.root, paths)
    found = sorted(projects.values(), key=lambda project: project.name.casefold())
    if PROJECT_FILTER:
        found = [project for project in found if project.name == PROJECT_FILTER]
    if PROJECT_START:
        start_index = next((index for index, project in enumerate(found) if project.name == PROJECT_START), None)
        if start_index is None:
            raise RuntimeError(f"ASIN_PROJECT_START project not found: {PROJECT_START}")
        found = found[start_index:]
    return found


def read_items(input_file: Path) -> list[dict[str, str]]:
    wb = load_workbook(input_file, data_only=True)
    ws = wb["ASIN清单"] if "ASIN清单" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [str(c.value or "").strip() for c in ws[1]]
    items = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: raw[i] if i < len(raw) else "" for i in range(len(headers))}
        if not any(row.values()):
            continue
        enabled = str(row.get("是否启用检查", "是") or "是").strip()
        if enabled and enabled not in {"是", "Y", "y", "yes", "YES", "1", "True", "TRUE"}:
            continue
        parent = str(row.get("父ASIN") or "").strip().upper()
        child = str(row.get("子ASIN") or row.get("ASIN") or "").strip().upper()
        sku = str(row.get("子SKU") or row.get("SKU") or "").strip()
        child_url = str(row.get("子ASIN网址") or row.get("链接") or row.get("URL") or "").strip()
        parent_url = str(row.get("父ASIN网址") or "").strip()
        if child and not child_url:
            child_url = f"https://www.amazon.com/dp/{child}"
        if parent and not parent_url:
            parent_url = f"https://www.amazon.com/dp/{parent}"
        if child:
            items.append(
                with_item_aliases(
                    {
                        "父ASIN": parent or child,
                        "子ASIN": child,
                        "子SKU": sku,
                        "子ASIN网址": child_url,
                        "父ASIN网址": parent_url,
                        "备注": str(row.get("备注") or ""),
                    }
                )
            )
    return items


def load_previous(paths: ProjectPaths) -> dict[str, Any]:
    if paths.snapshot_file.exists():
        return json.loads(paths.snapshot_file.read_text(encoding="utf-8"))
    return {}


def first_project_url(paths: ProjectPaths) -> str:
    items = read_items(paths.input_file)
    for item in items:
        url = str(item.get("子ASIN网址") or "").strip()
        if url:
            return url
    return "https://www.amazon.com/"


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(path)


def write_root_summary(payload: Any) -> None:
    write_json(LEGACY_ROOT / "last_run_summary.json", payload)
    if ROOT != LEGACY_ROOT:
        write_json(ROOT / "last_run_summary.json", payload)


def classify_precheck_failure(exc: Exception) -> tuple[str, str]:
    message = str(exc)
    if "ERR_NETWORK_ACCESS_DENIED" in message:
        return "network_access_denied", "Run outside sandbox or allow Amazon access before full crawl."
    if re.search(r"CAPTCHA|Enter the characters you see|Sorry, we just need", message, flags=re.I):
        return "captcha_or_risk_control", "Use a stable browser/network environment, then rerun the full crawl."
    return "precheck_failed", "Run outside sandbox or allow Amazon access before full crawl."


def run_mode_label(is_filtered_run: bool) -> str:
    return "filtered" if is_filtered_run else "full"


def curl_fetch_html(url: str, timeout: int = 40) -> str:
    curl_bin = shutil.which("curl.exe") or shutil.which("curl") or "curl.exe"
    cmd = [
        curl_bin,
        "-s",
        "-L",
        "--max-time",
        str(timeout),
        "-A",
        CURL_USER_AGENT,
        "-H",
        f"Accept: {CURL_ACCEPT}",
        "-H",
        "Accept-Language: en-US,en;q=0.9",
        url,
    ]
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="ignore",
            timeout=timeout + 5,
            check=False,
        )
    except Exception:
        return ""
    return result.stdout or ""


def amazon_access_precheck(sample_url: str) -> None:
    if sync_playwright is None:
        raise RuntimeError("Playwright 未安装，无法执行巡检预检。")
    goto_error = ""
    curl_html = ""
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(locale="en-US", viewport={"width": 1280, "height": 900})
        page = context.new_page()
        page.set_default_timeout(15000)
        page.set_default_navigation_timeout(30000)
        try:
            page.goto(sample_url, wait_until="domcontentloaded", timeout=30000)
            body_text = clean_text(page.locator("body").inner_text(timeout=10000))
            if re.search(r"captcha|Enter the characters you see|Sorry, we just need", body_text, flags=re.I):
                raise RuntimeError(f"Amazon precheck blocked by CAPTCHA at {sample_url}")
        except Exception as exc:
            goto_error = str(exc)
        finally:
            try:
                page.close()
            except Exception:
                pass
            context.close()
            browser.close()
    if not goto_error:
        return
    curl_html = curl_fetch_html(sample_url, timeout=20)
    if curl_html and re.search(r"<title|productTitle|acrCustomerReviewText|Buy Now|Add to Cart", curl_html, flags=re.I):
        raise RuntimeError(
            f"Amazon precheck failed in Playwright but curl still returns page HTML. "
            f"Current run should stop because screenshots/DOM capture would be unreliable. "
            f"sample={sample_url} error={goto_error}"
        )
    raise RuntimeError(
        f"Amazon precheck failed before crawl start. "
        f"Playwright cannot open Amazon and curl fallback did not return usable HTML. "
        f"sample={sample_url} error={goto_error}"
    )


def html_windows(raw_html: str, markers: list[str], size: int = 4500, limit: int = 3) -> str:
    if not raw_html:
        return ""
    parts = []
    for marker in markers:
        for match in re.finditer(marker, raw_html, flags=re.I):
            start = max(0, match.start() - 600)
            snippet = raw_html[start: match.start() + size]
            snippet = clean_text(snippet)
            if snippet and snippet not in parts:
                parts.append(snippet)
            if len(parts) >= limit:
                return "\n".join(parts)
    return "\n".join(parts)


def html_first_block(raw_html: str, selector_id: str, size: int = 5000) -> str:
    pattern = rf'id=["\']{re.escape(selector_id)}["\']'
    return html_windows(raw_html, [pattern], size=size, limit=1)


def html_detect_breadcrumb(raw_html: str) -> str:
    for marker in [
        r'id=["\']wayfinding-breadcrumbs_feature_div["\']',
        r'id=["\']wayfinding-breadcrumbs_container["\']',
        r'class=["\'][^"\']*\ba-breadcrumb\b',
    ]:
        match = re.search(marker, raw_html, flags=re.I)
        if not match:
            continue
        snippet = raw_html[match.start(): match.start() + 8000]
        parts = []
        for text in re.findall(r"<a\b[^>]*>(.*?)</a>", snippet, flags=re.I | re.S):
            cleaned = clean_text(text)
            if cleaned and cleaned not in parts:
                parts.append(cleaned)
        if parts:
            return " 鈥? ".join(parts)
    return ""


def html_detect_price(raw_html: str) -> dict[str, str]:
    price_scope_html = "\n".join(
        part
        for part in [
            html_first_block(raw_html, "corePriceDisplay_desktop_feature_div"),
            html_first_block(raw_html, "corePrice_feature_div"),
            html_first_block(raw_html, "apex_desktop"),
            html_first_block(raw_html, "desktop_buybox"),
            html_first_block(raw_html, "buybox"),
        ]
        if part
    )
    price_scope = clean_text(price_scope_html)
    list_price = ""
    typical_price = ""
    list_match = re.search(r"(?:List Price|Was:)\s*(\$[\d,]+(?:\.\d{2})?)", price_scope, flags=re.I)
    if list_match:
        list_price = list_match.group(1)
    typical_match = re.search(r"Typical Price\s*(\$[\d,]+(?:\.\d{2})?)", price_scope, flags=re.I)
    if typical_match:
        typical_price = typical_match.group(1)
    regular_match = re.search(r"Regular Price\s*:?[\s\S]{0,80}?(\$[\d,]+(?:\.\d{2})?)", price_scope, flags=re.I)
    regular_price = regular_match.group(1) if regular_match else ""
    prime_match = re.search(r"Prime Member Price\s*:?[\s\S]{0,80}?(\$[\d,]+(?:\.\d{2})?)", price_scope, flags=re.I)
    if not prime_match:
        prime_match = re.search(r"(\$[\d,]+(?:\.\d{2})?)[\s\S]{0,100}?(?:Exclusive Prime price|exclusively for Amazon Prime members)", price_scope, flags=re.I)
    prime_price = prime_match.group(1) if prime_match else ""
    current_price = ""
    pair_match = re.search(
        r'class=["\'][^"\']*a-price-whole[^"\']*["\'][^>]*>\s*([^<]+)\s*<.*?class=["\'][^"\']*a-price-fraction[^"\']*["\'][^>]*>\s*([^<]+)\s*<',
        price_scope_html,
        flags=re.I | re.S,
    )
    if pair_match:
        current_price = f"${clean_text(pair_match.group(1)).rstrip('.')}.{clean_text(pair_match.group(2))}"
    if not current_price:
        current_match = re.search(r"\$[\d,]+(?:\.\d{2})?", price_scope)
        if current_match:
            current_price = current_match.group(0)
    if regular_price:
        current_price = regular_price
    # Typical Price is informational; only List Price/Was is a strike-through basis.
    strike_basis = list_price
    discount = ""
    pct_match = re.search(r"(-\s*\d+%|\d+%\s*off)", price_scope, flags=re.I)
    if strike_basis and pct_match:
        discount = clean_text(pct_match.group(1))
    elif strike_basis and current_price:
        basis = parse_money_value(strike_basis)
        current = parse_money_value(current_price)
        if basis > current > 0:
            discount = f"-{round((basis - current) / basis * 100):.0f}%"
    return {
        "list_price": list_price,
        "typical_price": typical_price,
        "has_strike": "是" if strike_basis else "否",
        "current_price": current_price,
        "prime_offer": f"Prime会员专享折扣：{prime_price}" if prime_price else "",
        "discount": discount,
        "price_context": price_scope[:500],
    }


def html_detect_buybox(raw_html: str) -> str:
    html_lower = raw_html.lower()
    if re.search(
        r"Currently unavailable|No featured offers available|We don't know when or if this item will be back in stock",
        raw_html,
        flags=re.I,
    ):
        return "不可售/无购物车"
    if 'id="add-to-cart-button"' in html_lower or "add-to-cart-button" in html_lower:
        return "有购物车"
    if 'id="buy-now-button"' in html_lower or "buy-now-button" in html_lower:
        return "有Buy Now"
    if re.search(r"See All Buying Options", raw_html, flags=re.I):
        return "无购物车/仅购买选项"
    if 'id="desktop_buybox"' in html_lower or 'id="buybox"' in html_lower:
        return "购物车区域可见/未确认"
    return "未知"


def html_detect_title(raw_html: str) -> str:
    match = re.search(r'id=["\']productTitle["\'][^>]*>(.*?)</', raw_html, flags=re.I | re.S)
    if match:
        return clean_text(match.group(1))
    title_match = re.search(r"<title[^>]*>(.*?)</title>", raw_html, flags=re.I | re.S)
    return clean_text(title_match.group(1)) if title_match else ""


def html_detect_bullets(raw_html: str) -> str:
    match = re.search(r'id=["\']feature-bullets["\'][^>]*>(.*?)</ul>', raw_html, flags=re.I | re.S)
    if not match:
        return ""
    bullets = []
    for text in re.findall(r"<li\b[^>]*>(.*?)</li>", match.group(1), flags=re.I | re.S):
        cleaned = clean_text(text)
        if cleaned and cleaned not in bullets:
            bullets.append(cleaned)
    return "\n\n".join(bullets[:8])


def html_detect_rating(raw_html: str) -> str:
    match = re.search(r"([\d.]+)\s+out of\s+5", raw_html, flags=re.I)
    return f"{match.group(1)} out of 5" if match else ""


def html_detect_reviews(raw_html: str) -> str:
    block = html_first_block(raw_html, "acrCustomerReviewText", size=600)
    if not block:
        block = html_windows(raw_html, [r'acrCustomerReviewText'], size=600, limit=1)
    match = re.search(r"([\d,]+)", block)
    return f"{match.group(1)} ratings" if match else ""


def html_detect_star_distribution(raw_html: str) -> dict[str, float]:
    scope = html_windows(
        raw_html,
        [r"histogramTable", r"rating-histogram", r"Customer reviews", r"global ratings"],
        size=8000,
        limit=4,
    )
    return parse_star_distribution(scope)


def html_detect_delivery(raw_html: str) -> str:
    scope = "\n".join(
        part
        for part in [
            html_first_block(raw_html, "mir-layout-DELIVERY_BLOCK-slot-PRIMARY_DELIVERY_MESSAGE_LARGE"),
            html_first_block(raw_html, "deliveryBlockMessage"),
            html_first_block(raw_html, "delivery-message"),
            html_first_block(raw_html, "glow-ingress-line2"),
        ]
        if part
    )
    if not scope:
        scope = html_windows(raw_html, [r"Deliver to", r"Ships to"], size=1200, limit=2)
    return clean_text(scope)[:200]


def html_detect_location(raw_html: str) -> tuple[str, str]:
    zip_match = re.search(r'"zipCode"\s*:\s*"([^"]+)"', raw_html)
    country_match = re.search(r'"countryCode"\s*:\s*"([^"]+)"', raw_html)
    return (country_match.group(1) if country_match else "", zip_match.group(1) if zip_match else "")


def html_detect_seller(raw_html: str) -> dict[str, str]:
    seller = {}
    sold_by = ""
    for pattern in [
        r'id=["\']sellerProfileTriggerId["\'][^>]*>(.*?)</a>',
        r"Sold by\s*([^.<]+)",
    ]:
        match = re.search(pattern, raw_html, flags=re.I | re.S)
        if match:
            sold_by = clean_text(match.group(1))
            if sold_by:
                break
    if sold_by:
        seller["sold_by"] = sold_by
    scope = html_windows(raw_html, [r"Ships from", r"tabular-buybox", r"merchant-info"], size=2400, limit=3)
    ships_match = re.search(r"Ships from\s*([^.<]+)", scope, flags=re.I)
    if ships_match:
        seller["ships_from"] = clean_text(ships_match.group(1))
    return seller


def html_detect_aplus(raw_html: str) -> str:
    for marker in [
        r'id=["\']aplus["\']',
        r'id=["\']aplus_feature_div["\']',
        r'id=["\']dpx-aplus-product-description_feature_div["\']',
        r'id=["\']productDescription["\']',
    ]:
        if re.search(marker, raw_html, flags=re.I):
            return clean_text(html_windows(raw_html, [marker], size=2500, limit=1))
    return ""


def html_has_breadcrumb_marker(raw_html: str) -> bool:
    return bool(
        re.search(
            r'wayfinding-breadcrumbs_feature_div|wayfinding-breadcrumbs_container|\ba-breadcrumb\b',
            raw_html,
            flags=re.I,
        )
    )


def html_has_rank_marker(raw_html: str) -> bool:
    return bool(re.search(r"Best Sellers Rank|#\s*[\d,]+\s+in", raw_html, flags=re.I))


def is_not_found_page(title: str, body_text: str) -> bool:
    text = clean_text(" ".join([title or "", body_text or ""]))
    if not text:
        return False
    return bool(
        re.search(
            r"Page Not Found|Sorry!? We couldn'?t find that page|The Dogs of Amazon|looking for something\?",
            text,
            flags=re.I,
        )
    )


def extract_product_from_html(raw_html: str, item: dict[str, str], paths: ProjectPaths) -> dict[str, Any]:
    title = html_detect_title(raw_html)
    blocked = bool(re.search(r"Click the button below to continue shopping|Continue shopping", raw_html, flags=re.I))
    captcha = bool(re.search(r"captcha|Enter the characters you see|Sorry, we just need", raw_html, flags=re.I)) or blocked
    price_details = html_detect_price(raw_html)
    buybox = html_detect_buybox(raw_html)
    rating = html_detect_rating(raw_html)
    reviews = html_detect_reviews(raw_html)
    star_dist = html_detect_star_distribution(raw_html)
    review_shot = create_review_card(
        paths.screenshot_dir / "reviews" / f"{safe_name(item_value(item, '子ASIN', 'unknown'))}_customer_reviews_card_{STAMP}.png",
        rating,
        reviews,
        star_dist,
    )
    aplus = html_detect_aplus(raw_html)
    category = html_detect_breadcrumb(raw_html)
    rank_text = html_windows(
        raw_html,
        [
            r"Best Sellers Rank",
            r'id=["\']productDetails_detailBullets_sections1["\']',
            r'id=["\']detailBulletsWrapper_feature_div["\']',
            r'id=["\']prodDetails["\']',
        ],
        size=9000,
        limit=4,
    )
    ranks = parse_rank_entries(rank_text)
    rank = "\n".join(f"#{entry['rank']:,} in {entry['category']}" for entry in ranks[:2])
    country_code, zip_code = html_detect_location(raw_html)
    other_scope = clean_text(raw_html[:120000])
    not_found = is_not_found_page(title, other_scope)
    return normalize_unavailable_offer_fields({
        "status": "ERROR" if not_found or not title else "CAPTCHA" if captcha else "OK",
        "captcha": False if not_found else captcha,
        "blocked": False if not_found else blocked,
        "not_found": not_found,
        "title": title,
        "bullets": html_detect_bullets(raw_html),
        "aplus": aplus,
        "aplus_visible": "是" if aplus else "否",
        "rating": rating,
        "reviews": reviews,
        "rank": rank,
        "category": category,
        "buybox": buybox,
        "list_price": price_details["list_price"],
        "typical_price": price_details["typical_price"],
        "has_strike": price_details["has_strike"],
        "current_price": price_details["current_price"],
        "discount": price_details["discount"],
        "prime": price_details.get("prime_offer") or concise_offer_text(html_windows(raw_html, [r"Prime Member Price", r"exclusive Prime"], size=1200, limit=2)),
        "coupon": coupon_clean(clean_text(html_windows(raw_html, [r"coupon"], size=1800, limit=2))),
        "multi_buy": multi_buy_clean(html_windows(
            raw_html,
            [r"Save\s+\d+%\s+on\s+\d+", r"Buy\s+any\s+\d+", r"Buy\s+\d+", r"Get\s+\d+%\s+off"],
            size=1800,
            limit=4,
        )),
        "other_sellers": other_sellers(clean_text(html_windows(raw_html, [r"New\s*\(\d+\)\s*from", r"Other sellers"], size=1800, limit=2))),
        "star_distribution": star_dist,
        "negative_estimate": negative_estimate(reviews, star_dist),
        "star_distribution_screenshot": review_shot,
        "price_screenshot": "",
        "price_context": price_details["price_context"],
        "body_sample": clean_text(raw_html[:4000]),
        "seller": html_detect_seller(raw_html),
        "location_country": country_code,
        "location_zip": zip_code,
        "error": "Amazon returned Page Not Found" if not_found else "",
        "html_breadcrumb_marker_present": html_has_breadcrumb_marker(raw_html),
        "html_rank_marker_present": html_has_rank_marker(raw_html),
        "fetch_method": "curl_html",
    })


def merge_product_fields(primary: dict[str, Any], fallback: dict[str, Any]) -> dict[str, Any]:
    if not fallback:
        return primary
    merged = dict(primary)
    for key in [
        "title",
        "bullets",
        "aplus",
        "rating",
        "reviews",
        "rank",
        "category",
        "list_price",
        "typical_price",
        "has_strike",
        "current_price",
        "discount",
        "prime",
        "coupon",
        "multi_buy",
        "other_sellers",
        "price_context",
        "body_sample",
        "location_country",
        "location_zip",
        "delivery_zip_checked",
    ]:
        if not merged.get(key) and fallback.get(key):
            merged[key] = fallback.get(key)
    for key in ["html_breadcrumb_marker_present", "html_rank_marker_present"]:
        if key not in merged or merged.get(key) is None:
            merged[key] = fallback.get(key)
    if not merged.get("star_distribution") and fallback.get("star_distribution"):
        merged["star_distribution"] = fallback["star_distribution"]
    if not merged.get("star_distribution_screenshot") and fallback.get("star_distribution_screenshot"):
        merged["star_distribution_screenshot"] = fallback["star_distribution_screenshot"]
    if merged.get("buybox") in {"", "鏈煡"} and fallback.get("buybox"):
        merged["buybox"] = fallback["buybox"]
    if merged.get("aplus_visible") != "是" and fallback.get("aplus_visible") == "是":
        merged["aplus_visible"] = "是"
    if merged.get("status") == "ERROR" and fallback.get("status") in {"OK", "CAPTCHA"}:
        merged["status"] = fallback["status"]
        merged["captcha"] = fallback.get("captcha", False)
        merged["blocked"] = fallback.get("blocked", False)
    if not merged.get("fetch_method"):
        merged["fetch_method"] = fallback.get("fetch_method", "")
    return merged


def first_text(page, selectors: list[str]) -> str:
    for selector in selectors:
        try:
            loc = page.locator(selector).first
            if loc.count() and loc.is_visible(timeout=1200):
                text = clean_text(loc.inner_text(timeout=2500))
                if text:
                    return text
        except Exception:
            continue
    return ""


def all_text(page, selectors: list[str], limit: int = 8) -> str:
    parts = []
    for selector in selectors:
        try:
            loc = page.locator(selector)
            count = min(loc.count(), limit)
            for idx in range(count):
                text = clean_text(loc.nth(idx).inner_text(timeout=1200))
                if text and text not in parts:
                    parts.append(text)
        except Exception:
            continue
    return "\n\n".join(parts)


def visible_text_containing(page, patterns: list[str], limit: int = 220) -> str:
    try:
        body = clean_text(page.locator("body").inner_text(timeout=5000))
    except Exception:
        return ""
    return text_containing(body, patterns, limit)


def text_containing(body: str, patterns: list[str], limit: int = 220) -> str:
    hits = []
    for pattern in patterns:
        for match in re.finditer(pattern, body, flags=re.I):
            snippet = body[max(0, match.start() - 80): match.start() + limit]
            snippet = re.split(r"(?i)(Sponsored|Customers who|Back to top|Customer reviews)", snippet)[0].strip()
            if snippet and snippet not in hits:
                hits.append(snippet)
    return "；".join(hits[:3])


def html_window(page, marker: str, size: int = 4500) -> str:
    try:
        content = page.content()
    except Exception:
        return ""
    match = re.search(marker, content, flags=re.I)
    if not match:
        return ""
    return content[match.start(): match.start() + size]


def load_detail_sections(page) -> None:
    try:
        page.evaluate(
            """
            async () => {
              const sleep = ms => new Promise(resolve => setTimeout(resolve, ms));
              const height = Math.max(document.body.scrollHeight, 1);
              for (const y of [0.25, 0.5, 0.75, 1.0]) {
                window.scrollTo(0, Math.floor(height * y));
                await sleep(650);
              }
            }
            """
        )
        page.wait_for_timeout(800)
    except Exception:
        pass


def visible_money_from_selectors(page, selectors: list[str]) -> str:
    try:
        value = page.evaluate(
            """
            (selectors) => {
              const visible = (el) => {
                const r = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                return r.width > 0 && r.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
              };
              const badContext = /(per month|monthly|approval|store card|count\\)|shipping|delivery|coupon|save|list price|typical price|was:)/i;
              for (const selector of selectors) {
                for (const root of Array.from(document.querySelectorAll(selector))) {
                  const nodes = Array.from(root.querySelectorAll('.a-price .a-offscreen, .priceToPay .a-offscreen, [id*="priceblock"]'));
                  for (const node of nodes) {
                    const text = (node.textContent || '').trim();
                    const around = (node.closest('div, span, td, tr')?.innerText || text).replace(/\\s+/g, ' ');
                    if (!/^\\$\\s*\\d/.test(text)) continue;
                    if (!visible(node) && !visible(node.closest('.a-price') || node)) continue;
                    if (badContext.test(around)) continue;
                    return text.replace(/\\s+/g, '');
                  }
                  const rootText = (root.innerText || '').replace(/\\s+/g, ' ');
                  const match = rootText.match(/\\$\\s*\\d[\\d,]*(?:\\.\\d{2})?/);
                  if (match && !badContext.test(rootText.slice(Math.max(0, match.index - 50), match.index + 120))) {
                    return match[0].replace(/\\s+/g, '');
                  }
                }
              }
              return '';
            }
            """,
            selectors,
        )
        return value or ""
    except Exception:
        return ""


def screenshot_locator(page, selectors: list[str], output_path: Path) -> str:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    for selector in selectors:
        try:
            loc = page.locator(selector).first
            if loc.count() == 0:
                continue
            loc.scroll_into_view_if_needed(timeout=5000)
            page.wait_for_timeout(600)
            loc.screenshot(path=str(output_path), timeout=12000)
            return str(output_path)
        except Exception:
            continue
    return ""


def screenshot_price(page, output_path: Path) -> str:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(700)
        page.screenshot(path=str(output_path), full_page=False, timeout=15000)
        return str(output_path)
    except Exception:
        return screenshot_locator(page, ["#dp-container", "#ppd", "#centerCol", "#desktop_buybox", "#buybox"], output_path)


def screenshot_customer_reviews_card(page, output_path: Path) -> str:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        for selector in ["table#histogramTable", "table[data-hook='histogram']", "div[data-hook='rating-histogram']", "#reviewsMedley", "#reviewSummary", "text=Customer reviews"]:
            try:
                loc = page.locator(selector).first
                if loc.count():
                    loc.scroll_into_view_if_needed(timeout=5000)
                    page.wait_for_timeout(1200)
                    break
            except Exception:
                continue
        page.wait_for_timeout(1200)
        clip = page.evaluate(
            """
            () => {
              const root = document.querySelector('#reviewsMedley') || document;
              const useful = [];
              const add = (el) => {
                if (!el) return;
                const r = el.getBoundingClientRect();
                if (r.width <= 0 || r.height <= 0) return;
                if (r.bottom < 0 || r.top > window.innerHeight) return;
                useful.push(r);
              };
              Array.from(root.querySelectorAll('h2, h3, span, div, table, a')).forEach(el => {
                const t = (el.innerText || '').replace(/\\s+/g, ' ').trim();
                const r = el.getBoundingClientRect();
                if (r.width > 760 || r.height > 220) return;
                if (/^Customer reviews$/i.test(t)) add(el);
                if (/out of 5/i.test(t) && r.width < 520) add(el);
                if (/global ratings?/i.test(t) && r.width < 520) add(el);
                if (/^[1-5] star\\s+\\d+%$/i.test(t) || /^[1-5] star$/i.test(t) || /^\\d+%$/.test(t)) add(el);
                if (/How customer reviews and ratings work/i.test(t)) add(el);
              });
              add(root.querySelector('#histogramTable'));
              add(root.querySelector('table[data-hook="histogram"]'));
              if (!useful.length) return null;
              const left = Math.min(...useful.map(r => r.left));
              const top = Math.min(...useful.map(r => r.top));
              const right = Math.max(...useful.map(r => r.right));
              const bottom = Math.max(...useful.map(r => r.bottom));
              const pad = 14;
              return {
                x: Math.max(0, left - pad),
                y: Math.max(0, top - pad),
                width: Math.min(window.innerWidth - Math.max(0, left - pad), right - left + pad * 2),
                height: Math.min(window.innerHeight - Math.max(0, top - pad), bottom - top + pad * 2)
              };
            }
            """
        )
        if clip and clip.get("width", 0) <= 820 and clip.get("height", 0) <= 760:
            page.screenshot(path=str(output_path), clip=clip, timeout=12000)
            return str(output_path)
    except Exception:
        pass
    return screenshot_locator(
        page,
        ["#reviewSummary", "div[data-hook='rating-histogram']", "table#histogramTable"],
        output_path,
    )


def parse_star_distribution(text: str) -> dict[str, float]:
    result = {}
    for star, pct in re.findall(r"([1-5])\s+star\s+(\d{1,3})%", text, flags=re.I):
        result[f"{star}星"] = float(pct)
    return result


def extract_star_distribution_from_page(page) -> dict[str, float]:
    try:
        for selector in ["table#histogramTable", "table[data-hook='histogram']", "div[data-hook='rating-histogram']", "#reviewsMedley"]:
            try:
                loc = page.locator(selector).first
                if loc.count():
                    loc.scroll_into_view_if_needed(timeout=5000)
                    page.wait_for_timeout(1200)
                    break
            except Exception:
                continue
        rows = page.evaluate(
            """
            () => {
              const nodes = Array.from(document.querySelectorAll(
                '#histogramTable tr, table[data-hook="histogram"] tr, a[href*="filterByStar"]'
              ));
              const out = [];
              for (const node of nodes) {
                const text = (node.innerText || '').replace(/\s+/g, ' ').trim();
                const starMatch = text.match(/([1-5])\s*star/i);
                let pctMatch = text.match(/(\d{1,3})%/);
                if (!pctMatch) {
                  const extras = [
                    node.getAttribute('aria-label') || '',
                    node.getAttribute('title') || '',
                    node.querySelector('[aria-label*="%"]')?.getAttribute('aria-label') || '',
                    node.querySelector('[title*="%"]')?.getAttribute('title') || '',
                  ].join(' ');
                  pctMatch = extras.match(/(\d{1,3})%/);
                }
                if (starMatch && pctMatch) {
                  out.push([starMatch[1], pctMatch[1]]);
                }
              }
              return out;
            }
            """
        )
        result = {}
        for star, pct in rows or []:
            result.update(parse_star_distribution(f"{star} star {pct}%"))
        return result
    except Exception:
        return {}


def negative_estimate(reviews: str, distribution: dict[str, float]) -> int:
    total = parse_int(reviews)
    pct = sum(distribution.get(key, 0.0) for key in ("1星", "2星", "3星")) / 100
    return round(total * pct) if total else 0


def create_review_card(output_path: Path, rating: str, reviews: str, distribution: dict[str, float]) -> str:
    if not distribution:
        return ""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    width, height = 520, 380
    image = PILImage.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    try:
        title_font = ImageFont.truetype("arialbd.ttf", 28)
        normal_font = ImageFont.truetype("arial.ttf", 18)
        small_font = ImageFont.truetype("arial.ttf", 16)
    except Exception:
        title_font = normal_font = small_font = ImageFont.load_default()
    rating_match = re.search(r"(\d+(?:\.\d+)?)", rating or "")
    rating_value = rating_match.group(1) if rating_match else ""
    review_count = parse_int(reviews)
    draw.text((28, 22), "Customer reviews", fill=(0, 0, 0), font=title_font)
    def star_points(cx: int, cy: int, outer: int = 10, inner: int = 4) -> list[tuple[float, float]]:
        import math
        points = []
        for idx in range(10):
            angle = math.radians(-90 + idx * 36)
            radius = outer if idx % 2 == 0 else inner
            points.append((cx + radius * math.cos(angle), cy + radius * math.sin(angle)))
        return points

    for idx in range(5):
        draw.polygon(star_points(40 + idx * 18, 75), fill=(255, 102, 0))
    if rating_value:
        draw.text((132, 64), f"{rating_value} out of 5", fill=(0, 0, 0), font=normal_font)
    if review_count:
        draw.text((28, 96), f"{review_count:,} global ratings", fill=(80, 80, 80), font=small_font)
    y = 135
    for star in [5, 4, 3, 2, 1]:
        pct = distribution.get(f"{star}星", 0)
        draw.text((28, y - 2), f"{star} star", fill=(0, 85, 150), font=small_font)
        draw.rounded_rectangle((112, y, 362, y + 22), radius=4, outline=(120, 120, 120), width=1, fill=(250, 250, 250))
        fill_w = int(250 * pct / 100)
        if fill_w:
            draw.rectangle((113, y + 1, 113 + fill_w, y + 21), fill=(255, 102, 0))
        draw.text((382, y - 2), f"{pct:.0f}%", fill=(0, 85, 150), font=small_font)
        y += 42
    draw.text((28, 342), "How customer reviews and ratings work", fill=(0, 85, 150), font=small_font)
    image.save(output_path)
    return str(output_path)


def parse_rank_entries(text: str) -> list[dict[str, Any]]:
    entries = []
    for rank, category in re.findall(r"#\s*([\d,]+)\s+in\s+([^#\n]+)", text or "", flags=re.I):
        entries.append({"rank": int(rank.replace(",", "")), "category": clean_text(category)[:120]})
    return entries[:2]


def extract_rank_text(page) -> str:
    load_detail_sections(page)
    for selector in ["#productDetails_feature_div", "#prodDetails", "#productDetails_detailBullets_sections1", "#detailBulletsWrapper_feature_div"]:
        try:
            loc = page.locator(selector).first
            if loc.count():
                loc.scroll_into_view_if_needed(timeout=3500)
                page.wait_for_timeout(1200)
                break
        except Exception:
            continue
    snippets = []
    for selector in ["#productDetails_detailBullets_sections1", "#detailBulletsWrapper_feature_div", "#productDetails_db_sections", "#prodDetails"]:
        text = first_text(page, [selector])
        if re.search(r"Best Sellers Rank|#\s*[\d,]+\s+in", text, flags=re.I):
            snippets.append(text)
    try:
        body = page.locator("body").inner_text(timeout=7000)
        lines = [line.strip() for line in body.splitlines() if line.strip()]
        for idx, line in enumerate(lines):
            if re.search(r"Best Sellers Rank", line, flags=re.I):
                tail = lines[idx: idx + 8]
                rank_lines = [entry for entry in tail if re.search(r"#\s*[\d,]+\s+in\s+", entry)]
                if rank_lines:
                    return "\n".join(rank_lines[:2])
                snippets.append(" ".join(tail))
    except Exception:
        pass
    snippets.append(visible_text_containing(page, [r"Best Sellers Rank", r"#\s*[\d,]+\s+in"], limit=1200))
    for marker in [r"Best Sellers Rank", r"SalesRank", r"detailBulletsWrapper_feature_div", r"productDetails_detailBullets_sections1"]:
        window = clean_text(html_window(page, marker, 6000))
        if window:
            snippets.append(window)
    combined = " ".join(snippets)
    ranks = re.findall(r"#\s*([\d,]+)\s+in\s+([^#|]+?)(?=(?:#\s*[\d,]+\s+in|Customer Reviews|Date First|Product Dimensions|$))", combined, flags=re.I)
    if ranks:
        return "\n".join(f"#{num} in {clean_text(cat)}" for num, cat in ranks[:2])
    return ""


def extract_breadcrumb_category(page) -> str:
    selectors = [
        "#wayfinding-breadcrumbs_feature_div ul.a-unordered-list a",
        "#wayfinding-breadcrumbs_container ul.a-unordered-list a",
        "#wayfinding-breadcrumbs_feature_div .a-link-normal",
        "#wayfinding-breadcrumbs_container .a-link-normal",
    ]
    parts = []
    for selector in selectors:
        try:
            loc = page.locator(selector)
            count = min(loc.count(), 12)
            for idx in range(count):
                text = clean_text(loc.nth(idx).inner_text(timeout=1200))
                if text and text not in parts:
                    parts.append(text)
            if parts:
                break
        except Exception:
            continue
    if parts:
        return " › ".join(parts)

    breadcrumb_text = first_text(page, ["#wayfinding-breadcrumbs_feature_div", "#wayfinding-breadcrumbs_container"])
    if breadcrumb_text:
        pieces = [clean_text(piece) for piece in re.split(r"›|>|/", breadcrumb_text) if clean_text(piece)]
        if pieces:
            return " › ".join(pieces)
    return ""


def extract_price_details(page) -> dict[str, str]:
    core = clean_text(first_text(page, ["#corePriceDisplay_desktop_feature_div", "#apex_desktop", "#corePrice_feature_div"]))
    price_scope = clean_text(first_text(page, ["#corePriceDisplay_desktop_feature_div", "#corePrice_feature_div", "#apex_desktop"]))
    buybox_scope = clean_text(first_text(page, ["#desktop_buybox", "#buybox"]))
    list_price = ""
    typical_price = ""
    for label, target in [(r"List Price|Was:", "list"), (r"Typical Price", "typical")]:
        match = re.search(rf"(?:{label})\s*:?\s*(\$[\d,]+(?:\.\d{{2}})?)", price_scope, flags=re.I)
        if match and target == "list":
            list_price = match.group(1)
        if match and target == "typical":
            typical_price = match.group(1)
    regular_match = re.search(r"Regular Price\s*:?\s*(\$[\d,]+(?:\.\d{2})?)", buybox_scope, flags=re.I)
    regular_price = regular_match.group(1) if regular_match else ""
    prime_match = re.search(r"Prime Member Price\s*:?\s*(\$[\d,]+(?:\.\d{2})?)", buybox_scope, flags=re.I)
    prime_price = prime_match.group(1) if prime_match else ""
    if not prime_price and re.search(r"Exclusive Prime price", core, flags=re.I):
        prime_price = first_money(core)
    current_price = regular_price or first_money(core)
    if current_price and re.search(rf"{re.escape(current_price)}\s*(?:/|per)\s*(?:count|item|oz|fl)", core, flags=re.I):
        current_price = ""
    if not current_price:
        current_price = visible_money_from_selectors(page, ["#corePriceDisplay_desktop_feature_div", "#corePrice_feature_div", "#apex_desktop", "#desktop_buybox", "#buybox"])
    if not current_price:
        current_price = money(first_text(page, ["#priceblock_ourprice", "#priceblock_dealprice"]))
    # Typical Price must not independently create a strike-through discount.
    strike_basis = list_price
    discount = ""
    if current_price and strike_basis:
        current = parse_int(current_price) + (float(re.search(r"\.(\d{2})", current_price).group(0)) if re.search(r"\.\d{2}", current_price) else 0)
        basis = parse_int(strike_basis) + (float(re.search(r"\.(\d{2})", strike_basis).group(0)) if re.search(r"\.\d{2}", strike_basis) else 0)
        if basis > current > 0:
            discount = f"-{round((basis - current) / basis * 100):.0f}%"
    if not discount and strike_basis:
        discount = clean_text(visible_text_containing(page, [r"-\s*\d+%", r"\d+%\s+off"], limit=120))
    has_strike = "是" if strike_basis else "否"
    return {
        "list_price": list_price,
        "typical_price": typical_price,
        "has_strike": has_strike,
        "current_price": current_price,
        "prime_offer": f"Prime会员专享折扣：{prime_price}" if prime_price else "",
        "discount": clean_text(discount),
        "price_context": core[:500],
    }


def detect_buybox(page, current_price: str) -> str:
    body = first_text(page, ["body"])
    unavailable = first_text(page, ["#availability", "#outOfStock"])
    if re.search(r"Currently unavailable|We don't know when or if this item will be back in stock|Temporarily out of stock", unavailable or body, flags=re.I):
        return "不可售/无购物车"
    for selector, label in [
        ("#add-to-cart-button", "有购物车"),
        ("#buy-now-button", "有Buy Now"),
        ("#submit.add-to-cart", "有购物车"),
    ]:
        try:
            loc = page.locator(selector).first
            if loc.count() and loc.is_visible(timeout=1200):
                return label
        except Exception:
            continue
    if re.search(r"See All Buying Options", body, flags=re.I):
        return "无购物车/仅购买选项"
    for selector in ["#desktop_buybox", "#buybox"]:
        try:
            loc = page.locator(selector).first
            if loc.count() and loc.is_visible(timeout=1200):
                return "购物车区域可见/未确认"
        except Exception:
            continue
    return "未知"


def set_delivery_zip(page, postal_code: str = DEFAULT_POSTAL_CODE, target_url: str = "") -> bool:
    """Set and verify a delivery ZIP; failed changes must not produce Buy Box results."""
    url = target_url or "https://www.amazon.com/"
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=65000)
        page.wait_for_timeout(1200)

        opened = False
        for selector in ["#nav-global-location-popover-link", "#contextualIngressPtLabel"]:
            try:
                locator = page.locator(selector).first
                if locator.count() and locator.is_visible(timeout=1200):
                    locator.click(timeout=5000)
                    opened = True
                    break
            except Exception:
                continue
        if not opened:
            return False

        page.wait_for_timeout(800)
        filled = False
        for selector in [
            "#GLUXZipUpdateInput",
            "#GLUXZipUpdateInput_0",
            "input[name='GLUXZipUpdateInput']",
            "input[autocomplete='postal-code']",
        ]:
            try:
                locator = page.locator(selector).first
                if locator.count() and locator.is_visible(timeout=1200):
                    locator.fill(postal_code, timeout=5000)
                    filled = True
                    break
            except Exception:
                continue
        if not filled:
            return False

        submitted = False
        for selector in [
            "#GLUXZipUpdate input[type='submit']",
            "input[aria-labelledby='GLUXZipUpdate-announce']",
            "#GLUXZipUpdate",
        ]:
            try:
                locator = page.locator(selector).first
                if locator.count() and locator.is_visible(timeout=1200):
                    locator.click(timeout=5000)
                    submitted = True
                    break
            except Exception:
                continue
        if not submitted:
            return False

        page.wait_for_timeout(1800)
        for selector in ["button[name='glowDoneButton']", "input[name='glowDoneButton']"]:
            try:
                locator = page.locator(selector).first
                if locator.count() and locator.is_visible(timeout=800):
                    locator.click(timeout=2500)
                    break
            except Exception:
                continue

        # Reload the product page so availability reflects the new delivery area.
        page.goto(url, wait_until="domcontentloaded", timeout=65000)
        location_selectors = [
            "#glow-ingress-line2",
            "#nav-global-location-popover-link",
            "#contextualIngressPtLabel",
        ]
        for attempt in range(4):
            page.wait_for_timeout(800)
            location = first_text(page, location_selectors)
            if postal_code in location:
                return True
            if attempt == 1:
                page.goto(url, wait_until="domcontentloaded", timeout=65000)
        page.goto(url, wait_until="domcontentloaded", timeout=65000)
        for _ in range(3):
            page.wait_for_timeout(800)
            if postal_code in first_text(page, location_selectors):
                return True
        return False
    except Exception:
        # Amazon can apply the ZIP while the submit click reports an interrupted navigation.
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=65000)
            for _ in range(3):
                page.wait_for_timeout(800)
                location = first_text(
                    page,
                    ["#glow-ingress-line2", "#nav-global-location-popover-link", "#contextualIngressPtLabel"],
                )
                if postal_code in location:
                    return True
        except Exception:
            pass
        return False


def dismiss_amazon_continue(page) -> bool:
    try:
        body = first_text(page, ["body"])
        if "Click the button below to continue shopping" not in body and "Continue shopping" not in body:
            return False
        for selector in ["input[type='submit']", "button:has-text('Continue shopping')", "text=Continue shopping"]:
            try:
                loc = page.locator(selector).first
                if loc.count():
                    loc.click(timeout=3000)
                    page.wait_for_timeout(1800)
                    return True
            except Exception:
                continue
    except Exception:
        return False
    return False


def probe_buybox(page, item: dict[str, str], navigate: bool = True) -> str:
    url = item_value(item, "子ASIN网址")
    last_error = None
    if navigate:
        for attempt in range(2):
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=45000)
                if dismiss_amazon_continue(page):
                    page.goto(url, wait_until="domcontentloaded", timeout=45000)
                last_error = None
                break
            except Exception as exc:
                last_error = exc
                page.wait_for_timeout(800 * (attempt + 1))
    if last_error:
        raise last_error
    page.wait_for_timeout(1800)
    return detect_buybox(page, "")


def extract_with_zip_fallback(page, item: dict[str, str], paths: ProjectPaths) -> dict[str, Any]:
    current = extract_product(page, item, paths)
    current["delivery_zip_checked"] = ""
    current["buybox_zip_dependent"] = False
    current["buybox_zip_success"] = ""
    if (
        current.get("status") in {"ERROR", "CAPTCHA"}
        or current.get("not_found")
        or not is_cart_lost(current.get("buybox"))
    ):
        return current

    browser = page.context.browser
    if browser is None:
        return current

    attempted: list[str] = []
    for postal_code in POSTAL_CODES:
        attempted.append(postal_code)
        retry_context = browser.new_context(locale="en-US", viewport={"width": 1440, "height": 1200})
        retry_page = retry_context.new_page()
        retry_page.set_default_timeout(25000)
        retry_page.set_default_navigation_timeout(65000)
        retry = None
        try:
            url = item_value(item, "子ASIN网址")
            if not set_delivery_zip(retry_page, postal_code, url):
                continue
            probed_buybox = probe_buybox(retry_page, item, navigate=False)
            if not is_cart_lost(probed_buybox):
                retry = extract_product(retry_page, item, paths)
                retry["location_country"] = "US"
                retry["location_zip"] = postal_code
        except Exception:
            retry = None
        finally:
            retry_context.close()
        if retry is None:
            continue
        retry["delivery_zip_checked"] = " / ".join(attempted)
        retry["buybox_zip_dependent"] = False
        retry["buybox_zip_success"] = ""
        if (
            retry.get("status") not in {"ERROR", "CAPTCHA"}
            and not retry.get("not_found")
            and not is_cart_lost(retry.get("buybox"))
        ):
            retry["buybox_zip_dependent"] = True
            retry["buybox_zip_success"] = postal_code
            return retry
    current["delivery_zip_checked"] = " / ".join(attempted)
    current["buybox_zip_dependent"] = False
    current["buybox_zip_success"] = ""
    return current


def rank_cell(current_rank: str, previous_rank: str, index: int) -> tuple[str, bool]:
    current = parse_rank_entries(current_rank)
    before = parse_rank_entries(previous_rank)
    if index >= len(current):
        return "", False
    base = f"#{current[index]['rank']:,}"
    if index >= len(before):
        return base, False
    delta = current[index]["rank"] - before[index]["rank"]
    direction = "下降" if delta > 0 else "上升" if delta < 0 else "不变"
    pct = abs(delta) / before[index]["rank"] if before[index]["rank"] else 0
    return f"{base}\n较上次：#{before[index]['rank']:,} {direction}{abs(delta):,}名（{pct:.1%}）", pct > RANK_RED_THRESHOLD


def coupon_clean(text: str) -> str:
    text = clean_text(text)
    matches = []
    for pattern, fmt in [
        (r"Apply\s+(\d+%)\s+coupon", "Apply {} coupon"),
        (r"Apply\s+(\$\d+(?:\.\d{2})?)\s+coupon", "Apply {} coupon"),
        (r"Save\s+(\d+%)\s+with coupon", "Apply {} coupon"),
    ]:
        for match in re.finditer(pattern, text, flags=re.I):
            value = fmt.format(match.group(1))
            if value not in matches:
                matches.append(value)
    return "；".join(matches)


def multi_buy_clean(text: str) -> str:
    text = clean_text(text)
    matches = []
    patterns = [
        r"Save\s+\d+%\s+on\s+\d+\s+select\s+item\(s\)",
        r"Save\s+\d+%\s+on\s+\d+\s+selected?\s+items?",
        r"Buy\s+any\s+\d+\s*,?\s*Save\s+\d+%",
        r"Buy\s+\d+[^；。]{0,80}?(?:Save|Get)\s+\d+%",
        r"Get\s+\d+%\s+off\s+(?:when|if)\s+you\s+buy\s+\d+",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.I):
            value = clean_text(match.group(0))
            if value and value.casefold() not in {existing.casefold() for existing in matches}:
                matches.append(value)
    return "；".join(matches[:6])


def other_sellers(text: str) -> str:
    text = clean_text(text)
    match = re.search(r"New\s*\((\d+)\)\s*from", text, flags=re.I)
    if match and int(match.group(1)) > 1:
        return f"疑似跟卖：New ({match.group(1)}) from"
    return "无明显跟卖"


def extract_product(page, item: dict[str, str], paths: ProjectPaths) -> dict[str, Any]:
    child = item_value(item, "子ASIN")
    url = item_value(item, "子ASIN网址")
    last_error = None
    curl_html = ""
    curl_current: dict[str, Any] = {}
    for attempt in range(3):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=65000)
            if dismiss_amazon_continue(page):
                page.goto(url, wait_until="domcontentloaded", timeout=65000)
            last_error = None
            break
        except Exception as exc:
            last_error = exc
            page.wait_for_timeout(1200 * (attempt + 1))
    if last_error:
        curl_html = curl_fetch_html(url)
        curl_current = extract_product_from_html(curl_html, item, paths) if curl_html else {}
        if curl_current.get("title"):
            curl_current["error"] = str(last_error)
            curl_current["fetch_method"] = "curl_fallback_after_goto_error"
            return curl_current
        raise last_error
    page.wait_for_timeout(3500)
    body_text = clean_text(page.locator("body").inner_text(timeout=10000))
    blocked = bool(re.search(r"Click the button below to continue shopping|Continue shopping", body_text, flags=re.I))
    if blocked and dismiss_amazon_continue(page):
        page.goto(url, wait_until="domcontentloaded", timeout=65000)
        page.wait_for_timeout(2500)
        body_text = clean_text(page.locator("body").inner_text(timeout=10000))
        blocked = bool(re.search(r"Click the button below to continue shopping|Continue shopping", body_text, flags=re.I))
    captcha = bool(re.search(r"captcha|Enter the characters you see|Sorry, we just need", body_text, flags=re.I)) or blocked
    title = first_text(page, ["#productTitle", "span#title", "h1"])
    not_found = is_not_found_page(title, body_text)
    bullets = all_text(page, ["#feature-bullets li span.a-list-item", "#feature-bullets li"], limit=8)
    rating = first_text(page, ["#acrPopover", "span[data-hook='rating-out-of-text']", "i[data-hook='average-star-rating']"])
    reviews = first_text(page, ["#acrCustomerReviewText", "[data-hook='total-review-count']"])
    load_detail_sections(page)
    aplus = first_text(page, [
        "#aplus",
        "#aplus_feature_div",
        "#dpx-aplus-product-description_feature_div",
        "#aplus3p_feature_div",
        "#productDescription",
        "#productDescription_feature_div",
        "#dpx-product-description_feature_div",
    ])
    rank = extract_rank_text(page)
    category = extract_breadcrumb_category(page)
    try:
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(700)
    except Exception:
        pass
    price_details = extract_price_details(page)
    current_price = price_details["current_price"]
    buybox = detect_buybox(page, current_price)
    detail_body_text = clean_text(page.locator("body").inner_text(timeout=10000))
    coupon = coupon_clean(text_containing(detail_body_text, [r"Apply\s+\d+%\s+coupon", r"Apply\s+\$\d+.*coupon", r"coupon"], limit=220))
    prime = price_details.get("prime_offer") or text_containing(detail_body_text, [r"Prime Member Price", r"Exclusive Prime price", r"exclusively for Amazon Prime members"], limit=220)
    multi_buy = multi_buy_clean(text_containing(
        detail_body_text,
        [r"Save\s+\d+%\s+on\s+\d+", r"Buy\s+any\s+\d+", r"Buy\s+\d+", r"Get\s+\d+%\s+off"],
        limit=320,
    ))
    sellers = other_sellers(text_containing(detail_body_text, [r"New\s*\(\d+\)\s*from", r"Other sellers"], limit=260))
    body_text = detail_body_text
    star_dist = parse_star_distribution(body_text)
    if not star_dist:
        star_dist = extract_star_distribution_from_page(page)
    price_shot = screenshot_price(page, paths.screenshot_dir / "price" / f"{safe_name(child)}_price_{STAMP}.png")
    review_shot = screenshot_customer_reviews_card(page, paths.screenshot_dir / "reviews" / f"{safe_name(child)}_customer_reviews_{STAMP}.png")
    if not star_dist:
        try:
            star_dist = parse_star_distribution(clean_text(page.locator("body").inner_text(timeout=5000)))
        except Exception:
            star_dist = {}
    if not star_dist:
        star_dist = extract_star_distribution_from_page(page)
    try:
        if review_shot:
            with PILImage.open(review_shot) as im:
                w, h = im.size
            if w < 300 or h < 220 or h > 900:
                review_shot = ""
    except Exception:
        review_shot = ""
    if not review_shot:
        review_shot = create_review_card(paths.screenshot_dir / "reviews" / f"{safe_name(child)}_customer_reviews_card_{STAMP}.png", rating, reviews, star_dist)
    current = {
        "status": "CAPTCHA" if captcha else "ERROR" if not_found or not title else "OK",
        "captcha": captcha,
        "blocked": blocked,
        "not_found": not_found,
        "title": title,
        "bullets": bullets,
        "aplus": aplus,
        "aplus_visible": "是" if aplus else "否",
        "rating": rating,
        "reviews": reviews,
        "rank": rank,
        "category": category,
        "buybox": buybox,
        "list_price": price_details["list_price"],
        "typical_price": price_details["typical_price"],
        "has_strike": price_details["has_strike"],
        "current_price": current_price,
        "discount": price_details["discount"],
        "prime": clean_text(prime),
        "coupon": coupon,
        "multi_buy": multi_buy,
        "other_sellers": sellers,
        "star_distribution": star_dist,
        "negative_estimate": negative_estimate(reviews, star_dist),
        "star_distribution_screenshot": review_shot,
        "price_screenshot": price_shot,
        "price_context": price_details["price_context"],
        "body_sample": body_text[:800],
        "error": "Amazon returned Page Not Found" if not_found else "",
        "fetch_method": "playwright",
    }
    page_html = ""
    html_current: dict[str, Any] = {}
    if not current.get("rank") or not current.get("category") or current.get("buybox") == "未知":
        try:
            page_html = page.content()
        except Exception:
            page_html = ""
        if page_html:
            html_current = extract_product_from_html(page_html, item, paths)
            current = merge_product_fields(current, html_current)
    need_html_fallback = (
        not current.get("title")
        or not current.get("rank")
        or not current.get("category")
        or not current.get("current_price")
        or current.get("buybox") == "未知"
        or current.get("status") == "CAPTCHA"
    )
    if need_html_fallback:
        curl_html = curl_html or curl_fetch_html(url)
        curl_current = extract_product_from_html(curl_html, item, paths) if curl_html else {}
        current = merge_product_fields(current, curl_current)
        if current.get("status") == "ERROR" and curl_current.get("status") == "OK" and not curl_current.get("not_found"):
            current["status"] = "OK"
            current["not_found"] = False
            if current.get("error") == "Amazon returned Page Not Found":
                current["error"] = ""
        if current.get("status") == "CAPTCHA" and curl_current.get("status") == "OK":
            current["status"] = "OK"
            current["captcha"] = False
            current["blocked"] = False
        if current.get("buybox") in {"", "未知"} and curl_current.get("buybox"):
            current["buybox"] = curl_current["buybox"]
        if curl_current.get("location_country") or curl_current.get("location_zip"):
            current["location_country"] = curl_current.get("location_country", "")
            current["location_zip"] = curl_current.get("location_zip", "")
    final_not_found = is_not_found_page(str(current.get("title", "")), str(current.get("body_sample", "")))
    if final_not_found:
        current["status"] = "ERROR"
        current["not_found"] = True
        current["captcha"] = False
        current["blocked"] = False
        current["error"] = "Amazon returned Page Not Found"
    return normalize_unavailable_offer_fields(current)


def copy_text(current: dict[str, Any]) -> str:
    bullets = "\n\n".join([line.strip() for line in re.split(r"\n+|•", current.get("bullets") or "") if line.strip()])
    parts = []
    if current.get("title"):
        parts.append(f"【标题】{current['title']}")
    if bullets:
        parts.append(f"【五点】{bullets}")
    parts.append(f"【A+】{current.get('aplus') or '未显示/未采集到'}")
    return "\n\n".join(parts)


def change_ratio(before_value: Any, current_value: Any) -> float:
    before = parse_money_value(before_value)
    current = parse_money_value(current_value)
    if not before or not current:
        return 0.0
    return (current - before) / before


def issue_summary_text(issues: list[dict[str, str]]) -> str:
    parts = []
    for issue in issues:
        text = f"{issue.get('问题模块', '')}：{issue.get('问题摘要', '')}".strip("：")
        if text and text not in parts:
            parts.append(text)
    return "\n\n".join(parts)


def readable_issue_notes(
    issues: list[dict[str, str]],
    change_notes: list[str],
    extra_notes: list[str],
) -> str:
    blocks = [block for block in issue_summary_text(issues).split("\n\n") if block]

    def normalized(value: str) -> str:
        return re.sub(r"[\s：:；;，,。.]", "", clean_text(value)).casefold()

    def append_unique(value: Any) -> None:
        text = clean_text(value)
        key = normalized(text)
        if not key:
            return
        existing = [normalized(block) for block in blocks]
        if any(key == current or key in current for current in existing):
            return
        blocks.append(text)

    for note_group in change_notes:
        for note in re.split(r"；+|\n+", str(note_group or "")):
            append_unique(note)
    for note in extra_notes:
        append_unique(note)
    return "\n\n".join(blocks)


def parent_rank_baselines(previous: dict[str, Any]) -> dict[str, str]:
    baselines: dict[str, str] = {}
    for row in previous.values():
        parent = row.get("parent")
        rank = row.get("rank")
        if parent and rank and parent not in baselines:
            baselines[parent] = rank
    return baselines


def missing_category_issue(current: dict[str, Any]) -> tuple[str, str, str, str]:
    marker = current.get("html_breadcrumb_marker_present")
    if marker is False:
        return (
            "P2",
            "字段缺失",
            "页面未显示顶部breadcrumb，类目节点无法采集。",
            "确认页面顶部breadcrumb是否真实缺失；如需业务复核，请切换稳定环境后重试。",
        )
    if marker is True:
        return (
            "P2",
            "字段缺失",
            "页面存在顶部breadcrumb标记，但类目节点未采集到。",
            "人工打开页面确认顶部breadcrumb是否显示，必要时补选择器。",
        )
    return (
        "P2",
        "字段缺失",
        "类目节点未采集到。",
        "人工打开页面确认顶部breadcrumb是否显示，必要时补选择器。",
    )


def missing_rank_issue(current: dict[str, Any], previous_rank: str) -> tuple[str, str, str, str]:
    marker = current.get("html_rank_marker_present")
    if previous_rank:
        if marker is False:
            return (
                "P2",
                "字段缺失",
                "上次有排名，本次页面未显示 Best Sellers Rank。",
                "确认前台是否真实未显示 Best Sellers Rank；如需业务复核，请切换稳定环境后重试。",
            )
        if marker is True:
            return (
                "P2",
                "字段缺失",
                "上次有排名，本次页面存在 Best Sellers Rank 标记但未采集到。",
                "人工打开页面确认 Best Sellers Rank 是否显示，必要时补选择器。",
            )
        return (
            "P2",
            "字段缺失",
            "上次有排名，本次排名未采集到。",
            "人工打开页面确认前台是否仍显示排名，必要时补选择器。",
        )
    if marker is False:
        return (
            "P2",
            "字段缺失",
            "页面未显示 Best Sellers Rank，大小类排名无法采集。",
            "确认前台是否真实未显示 Best Sellers Rank；如需业务复核，请切换稳定环境后重试。",
        )
    if marker is True:
        return (
            "P2",
            "字段缺失",
            "页面存在 Best Sellers Rank 标记，但大小类排名未采集到。",
            "人工打开页面确认 Best Sellers Rank 是否显示，必要时补选择器。",
        )
    return (
        "P2",
        "字段缺失",
        "大小类排名未采集到。",
        "人工打开页面确认前台是否显示 Best Sellers Rank，必要时补选择器。",
    )


def compare(item: dict[str, str], current: dict[str, Any], previous: dict[str, Any]) -> tuple[list[dict[str, str]], str]:
    child = item_value(item, "子ASIN")
    before = previous.get(child, {})
    previous_rank = before.get("rank", "")
    current_rank = current.get("rank", "")
    issues = []
    notes = []
    if current.get("captcha"):
        issues.append(("P1", "风控/验证码", "Amazon 返回风控或验证码，本次数据不可作为可靠基准。", "切换稳定浏览器环境或人工复核后重跑。"))
    for label, field in [("标题", "title"), ("评分星级", "rating"), ("评论数", "reviews")]:
        if not current.get(field) and not current.get("captcha"):
            issues.append(("P2", "字段缺失", f"{label}未采集到。", "人工打开页面确认前台是否显示，必要时补选择器。"))
    if not current_rank and not current.get("captcha"):
        if previous_rank:
            issues.append(("P2", "字段缺失", "上次有排名，本次排名未采集到。", "人工打开页面确认前台是否仍显示排名，必要时补选择器。"))
        else:
            issues.append(("P2", "字段缺失", "大小类排名未采集到。", "人工打开页面确认前台是否显示 Best Sellers Rank，必要时补选择器。"))
    if current.get("parent_review_split"):
        issues.append(("P1", "父体评论拆分", current["parent_review_split"], "核对该父体变体是否存在评论拆分或异常合并。"))
        notes.append(current["parent_review_split"])
    if before and not current.get("captcha"):
        review_delta = parse_int(current.get("reviews")) - parse_int(before.get("reviews"))
        current["review_delta"] = review_delta
        if review_delta:
            notes.append(f"评论数变化：{parse_int(before.get('reviews')):,} -> {parse_int(current.get('reviews')):,}")
        before_negative = before.get("negative_estimate")
        before_negative_is_valid = before_negative not in (None, "", 0, "0")
        if before_negative_is_valid and current.get("negative_estimate", 0) > int(before_negative or 0):
            issues.append(("P1", "新增差评", f"估算差评数增加：{before.get('negative_estimate', 0)} -> {current.get('negative_estimate')}", "优先排查新增1-3星评价内容。"))
            notes.append(f"新来差评：估算差评数 {before.get('negative_estimate', 0)} -> {current.get('negative_estimate')}")
        if before.get("current_price") and not current.get("current_price"):
            issues.append(("P2", "字段缺失", "上次有Buy Box价格，本次价格未采集到。", "查看价格截图确认前台是否隐藏价格或选择器失效。"))
        price_delta = change_ratio(before.get("current_price"), current.get("current_price"))
        if abs(price_delta) > 0.10:
            direction = "上浮" if price_delta > 0 else "下降"
            issues.append(("P2", "价格变化", f"Buy Box价格{direction}超过10%：{before.get('current_price')} -> {current.get('current_price')}（{price_delta:+.1%}）", "核对促销、Coupon和购物车价格。"))
        current_ranks = parse_rank_entries(current_rank)
        before_ranks = parse_rank_entries(previous_rank)
        for rank_index, rank_label in [(0, "大类排名"), (1, "小类排名")]:
            if rank_index >= len(current_ranks) or rank_index >= len(before_ranks):
                continue
            old_rank = before_ranks[rank_index]["rank"]
            new_rank = current_ranks[rank_index]["rank"]
            if not old_rank:
                continue
            rank_delta = new_rank - old_rank
            rank_pct = abs(rank_delta) / old_rank
            if rank_pct > RANK_RED_THRESHOLD:
                direction = "下降" if rank_delta > 0 else "上升"
                summary = f"{rank_label}{direction}超过10%：#{old_rank:,} -> #{new_rank:,}（{rank_pct:.1%}）"
                issues.append(("P2", "排名变化", summary, "核对自然排名变化是否由广告、断货、价格或转化波动导致。"))
                notes.append(summary)
        if before.get("title") and current.get("title") and before.get("title") != current.get("title"):
            issues.append(("P2", "文案变化", "标题与昨日基准不一致。", "确认是否为计划内改版。"))
        if before.get("aplus_visible") == "是" and current.get("aplus_visible") != "是":
            issues.append(("P2", "A+变化", "上次显示A+，本次未显示或未采集到。", "打开页面确认A+模块是否真实丢失。"))
        if before.get("buybox") and not is_cart_lost(before.get("buybox")) and is_cart_lost(current.get("buybox")):
            issues.append(("P1", "购物车丢失", f"上次有购物车，本次为：{current.get('buybox') or '未知'}", "切换邮编复核购物车和Buy Now区域。"))
    if current.get("other_sellers", "").startswith("疑似跟卖"):
        issues.append(("P1", "跟卖", current["other_sellers"], "检查报价卖家并发起跟卖处理流程。"))
    rows = []
    for priority, module, summary, action in issues:
        rows.append({
            "日期": DATE,
            "父ASIN": item_value(item, "父ASIN"),
            "子ASIN": child,
            "健康状态": "异常",
            "最高优先级": priority,
            "问题模块": module,
            "是否需处理": "是",
            "问题摘要": summary,
            "关键变化": "；".join(notes),
            "证据/截图": "\n".join([p for p in [current.get("star_distribution_screenshot"), current.get("price_screenshot")] if p]),
            "建议动作": action,
            "责任跟进项": "",
            "处理状态": "未处理",
            "备注": item_value(item, "备注"),
        })
    return rows, "；".join(notes)


def compare(item: dict[str, str], current: dict[str, Any], previous: dict[str, Any]) -> tuple[list[dict[str, str]], str]:
    child = item_value(item, "子ASIN")
    before = previous.get(child, {})
    previous_rank = before.get("rank", "")
    current_rank = current.get("rank", "")
    issues = []
    notes = []
    if current.get("captcha"):
        issues.append(("P1", "风控/验证码", "Amazon 返回风控或验证码，本次数据不可作为可靠基准。", "切换稳定浏览器环境或人工复核后重跑。"))
    for label, field in [("标题", "title"), ("评分星级", "rating"), ("评论数", "reviews")]:
        if not current.get(field) and not current.get("captcha"):
            issues.append(("P2", "字段缺失", f"{label}未采集到。", "人工打开页面确认前台是否显示，必要时补选择器。"))
    if not current.get("category") and not current.get("captcha"):
        issues.append(missing_category_issue(current))
    if not current_rank and not current.get("captcha"):
        issues.append(missing_rank_issue(current, previous_rank))
    if current.get("parent_review_split"):
        issues.append(("P1", "父体评论拆分", current["parent_review_split"], "核对该父体变体是否存在评论拆分或异常合并。"))
        notes.append(current["parent_review_split"])
    if before and not current.get("captcha"):
        review_delta = parse_int(current.get("reviews")) - parse_int(before.get("reviews"))
        current["review_delta"] = review_delta
        if review_delta:
            notes.append(f"评论数变化：{parse_int(before.get('reviews')):,} -> {parse_int(current.get('reviews')):,}")
        before_negative = before.get("negative_estimate")
        before_negative_is_valid = before_negative not in (None, "", 0, "0")
        if before_negative_is_valid and current.get("negative_estimate", 0) > int(before_negative or 0):
            issues.append(("P1", "新增差评", f"估算差评数增加：{before.get('negative_estimate', 0)} -> {current.get('negative_estimate')}", "优先排查新增1-3星评价内容。"))
            notes.append(f"新来差评：估算差评数 {before.get('negative_estimate', 0)} -> {current.get('negative_estimate')}")
        if before.get("current_price") and not current.get("current_price"):
            issues.append(("P2", "字段缺失", "上次有Buy Box价格，本次价格未采集到。", "查看价格截图确认前台是否隐藏价格或选择器失效。"))
        price_delta = change_ratio(before.get("current_price"), current.get("current_price"))
        if abs(price_delta) > 0.10:
            direction = "上浮" if price_delta > 0 else "下降"
            issues.append(("P2", "价格变化", f"Buy Box价格{direction}超过10%：{before.get('current_price')} -> {current.get('current_price')}（{price_delta:+.1%}）", "核对促销、Coupon和购物车价格。"))
        current_ranks = parse_rank_entries(current_rank)
        before_ranks = parse_rank_entries(previous_rank)
        for rank_index, rank_label in [(0, "大类排名"), (1, "小类排名")]:
            if rank_index >= len(current_ranks) or rank_index >= len(before_ranks):
                continue
            old_rank = before_ranks[rank_index]["rank"]
            new_rank = current_ranks[rank_index]["rank"]
            if not old_rank:
                continue
            rank_delta = new_rank - old_rank
            rank_pct = abs(rank_delta) / old_rank
            if rank_pct > RANK_RED_THRESHOLD:
                direction = "下降" if rank_delta > 0 else "上升"
                summary = f"{rank_label}{direction}超过10%：#{old_rank:,} -> #{new_rank:,}（{rank_pct:.1%}）"
                issues.append(("P2", "排名变化", summary, "核对自然排名变化是否由广告、断货、价格或转化波动导致。"))
                notes.append(summary)
        if before.get("title") and current.get("title") and before.get("title") != current.get("title"):
            issues.append(("P2", "文案变化", "标题与昨日基准不一致。", "确认是否为计划内改版。"))
        if before.get("aplus_visible") == "是" and current.get("aplus_visible") != "是":
            issues.append(("P2", "A+变化", "上次显示A+，本次未显示或未采集到。", "打开页面确认A+模块是否真实丢失。"))
        if before.get("buybox") and not is_cart_lost(before.get("buybox")) and is_cart_lost(current.get("buybox")):
            issues.append(("P1", "购物车丢失", f"上次有购物车，本次为：{current.get('buybox') or '未知'}", "切换邮编复核购物车和Buy Now区域。"))
    if current.get("other_sellers", "").startswith("疑似跟卖"):
        issues.append(("P1", "跟卖", current["other_sellers"], "检查报价卖家并发起跟卖处理流程。"))
    rows = []
    for priority, module, summary, action in issues:
        rows.append({
            "日期": DATE,
            "父ASIN": item_value(item, "父ASIN"),
            "子ASIN": child,
            "健康状态": "异常",
            "最高优先级": priority,
            "问题模块": module,
            "是否需处理": "是",
            "问题摘要": summary,
            "关键变化": "；".join(notes),
            "证据/截图": "\n".join([p for p in [current.get("star_distribution_screenshot"), current.get("price_screenshot")] if p]),
            "建议动作": action,
            "责任跟进人": "",
            "处理状态": "未处理",
            "备注": item_value(item, "备注"),
        })
    return rows, "；".join(notes)


def compare(item: dict[str, str], current: dict[str, Any], previous: dict[str, Any]) -> tuple[list[dict[str, str]], str]:
    child = item_value(item, "子ASIN")
    before = previous.get(child, {})
    previous_rank = before.get("rank", "")
    current_rank = current.get("rank", "")
    issues = []
    notes = []

    if current.get("not_found"):
        issues.append((
            "P1",
            "链接失效",
            f"子ASIN {child} 返回 Page Not Found。",
            "确认 ASIN/链接是否失效、下架或跳转错误；必要时更新输入表链接。",
        ))
    else:
        if current.get("captcha"):
            issues.append(("P1", "风控/验证码", "Amazon 返回风控或验证码，本次数据不可作为可靠基准。", "切换稳定浏览器环境或人工复核后重跑。"))
        for label, field in [("标题", "title"), ("评分星级", "rating"), ("评论数", "reviews")]:
            if not current.get(field) and not current.get("captcha"):
                issues.append(("P2", "字段缺失", f"{label}未采集到。", "人工打开页面确认前台是否显示，必要时补选择器。"))
        if not current.get("category") and not current.get("captcha"):
            issues.append(missing_category_issue(current))
        if not current_rank and not current.get("captcha"):
            issues.append(missing_rank_issue(current, previous_rank))
        if current.get("parent_review_split"):
            issues.append(("P1", "父体评论拆分", current["parent_review_split"], "核对该父体变体是否存在评论拆分或异常合并。"))

    if before and not current.get("captcha") and not current.get("not_found"):
        review_delta = parse_int(current.get("reviews")) - parse_int(before.get("reviews"))
        current["review_delta"] = review_delta
        if review_delta:
            notes.append(f"评论数变化：{parse_int(before.get('reviews')):,} -> {parse_int(current.get('reviews')):,}")
        before_negative = before.get("negative_estimate")
        before_negative_is_valid = before_negative not in (None, "", 0, "0")
        if before_negative_is_valid and current.get("negative_estimate", 0) > int(before_negative or 0):
            issues.append(("P1", "新增差评", f"估算差评数增加：{before.get('negative_estimate', 0)} -> {current.get('negative_estimate')}", "优先排查新增1-3星评价内容。"))
            notes.append(f"新来差评：估算差评数 {before.get('negative_estimate', 0)} -> {current.get('negative_estimate')}")
        if before.get("current_price") and not current.get("current_price"):
            issues.append(("P2", "字段缺失", "上次有Buy Box价格，本次价格未采集到。", "查看价格截图确认前台是否隐藏价格或选择器失效。"))
        price_delta = change_ratio(before.get("current_price"), current.get("current_price"))
        if abs(price_delta) > 0.10:
            direction = "上浮" if price_delta > 0 else "下降"
            issues.append(("P2", "价格变化", f"Buy Box价格{direction}超过10%：{before.get('current_price')} -> {current.get('current_price')}（{price_delta:+.1%}）", "核对促销、Coupon和购物车价格。"))
        current_ranks = parse_rank_entries(current_rank)
        before_ranks = parse_rank_entries(previous_rank)
        for rank_index, rank_label in [(0, "大类排名"), (1, "小类排名")]:
            if rank_index >= len(current_ranks) or rank_index >= len(before_ranks):
                continue
            old_rank = before_ranks[rank_index]["rank"]
            new_rank = current_ranks[rank_index]["rank"]
            if not old_rank:
                continue
            rank_delta = new_rank - old_rank
            rank_pct = abs(rank_delta) / old_rank
            if rank_pct > RANK_RED_THRESHOLD:
                direction = "下降" if rank_delta > 0 else "上升"
                summary = f"{rank_label}{direction}超过10%：#{old_rank:,} -> #{new_rank:,}（{rank_pct:.1%}）"
                issues.append(("P2", "排名变化", summary, "核对自然排名变化是否由广告、断货、价格或转化波动导致。"))
                notes.append(summary)
        if before.get("title") and current.get("title") and before.get("title") != current.get("title"):
            issues.append(("P2", "文案变化", "标题与昨日基准不一致。", "确认是否为计划内改版。"))
        if before.get("aplus_visible") == "是" and current.get("aplus_visible") != "是":
            issues.append(("P2", "A+变化", "上次显示A+，本次未显示或未采集到。", "打开页面确认A+模块是否真实丢失。"))

    if not current.get("captcha") and not current.get("not_found") and is_cart_lost(current.get("buybox")):
        if before.get("buybox") and not is_cart_lost(before.get("buybox")):
            cart_summary = f"上次有购物车，本次为：{current.get('buybox') or '未知'}"
        else:
            cart_summary = f"本次未检测到 Add to Cart/Buy Now：{current.get('buybox') or '未知'}"
        issues.append(("P1", "购物车丢失", cart_summary, "切换邮编复核购物车和Buy Now区域。"))

    if current.get("other_sellers", "").startswith("疑似跟卖"):
        issues.append(("P1", "跟卖", current["other_sellers"], "检查报价卖家并发起跟卖处理流程。"))

    rows = []
    for priority, module, summary, action in issues:
        rows.append({
            "日期": DATE,
            "父ASIN": item_value(item, "父ASIN"),
            "子ASIN": child,
            "健康状态": "异常",
            "最高优先级": priority,
            "问题模块": module,
            "是否需处理": "是",
            "问题摘要": summary,
            "关键变化": "；".join(notes),
            "证据/截图": "\n".join([p for p in [current.get("star_distribution_screenshot"), current.get("price_screenshot")] if p]),
            "建议动作": action,
            "责任跟进人": "",
            "处理状态": "未处理",
            "备注": item_value(item, "备注"),
        })
    return rows, "；".join(notes)


def unique_sheet_name(wb: Workbook, base: str) -> str:
    if base not in wb.sheetnames:
        return base
    idx = 1
    while f"{base}-{idx}" in wb.sheetnames:
        idx += 1
    return f"{base}-{idx}"


def ensure_book(path: Path) -> Workbook:
    return load_workbook(path) if path.exists() else Workbook()


def reset_default_sheet(wb: Workbook) -> None:
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]


def write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"


def style_ws(ws) -> None:
    widths = {
        "日期/检查内容": 18, "父ASIN": 14, "子SKU": 20, "子ASIN": 14, "是否异常": 10,
        "购物车": 18, "文案": 55, "A+是否显示": 12, "评分星级": 16, "评论数": 12,
        "星级占比截图": 52, "跟卖情况": 24, "类目节点": 28, "大类排名": 25, "小类排名": 25,
        "价格截图": 76, "备注": 28, "问题摘要": 45, "关键变化": 35, "证据/截图": 45, "建议动作": 40,
    }
    header_to_col = {cell.value: cell.column for cell in ws[1]}
    for header, width in widths.items():
        if header in header_to_col:
            ws.column_dimensions[get_column_letter(header_to_col[header])].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 18, 42)
    for header in ("大类排名", "小类排名"):
        col = header_to_col.get(header)
        if not col:
            continue
        for row_idx in range(2, ws.max_row + 1):
            match = re.search(r"（([\d.]+)%）", str(ws.cell(row_idx, col).value or ""))
            if match and float(match.group(1)) / 100 > RANK_RED_THRESHOLD:
                ws.cell(row_idx, col).font = Font(color="FF0000")


def embed_images(ws, paths: ProjectPaths) -> None:
    ws._images = []
    header_to_col = {cell.value: cell.column for cell in ws[1]}
    cfgs = {"星级占比截图": 360, "价格截图": 560}
    embed_dir = paths.cache_dir / "excel_embedded" / DATE
    embed_dir.mkdir(parents=True, exist_ok=True)
    for header, max_width in cfgs.items():
        col = header_to_col.get(header)
        if not col:
            continue
        col_letter = get_column_letter(col)
        for row_idx in range(2, ws.max_row + 1):
            source = ws.cell(row_idx, col).value
            if not source or not Path(str(source)).exists():
                continue
            source_path = Path(str(source))
            try:
                with PILImage.open(source_path) as original:
                    width, height = original.size
                scale = min(1.0, max_width / width) if width else 1.0
                display_w = int(width * scale)
                display_h = int(height * scale)
                img = XLImage(str(source_path))
                img.width = display_w
                img.height = display_h
                img.object_position = 1
                ws.add_image(img, f"{col_letter}{row_idx}")
                ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 0, display_h * 0.75 + 12)
            except Exception:
                continue


def save_book(wb: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_updated_{STAMP}{path.suffix}")
        wb.save(fallback)
        return fallback


def new_output_path(paths: ProjectPaths, is_filtered_run: bool = False) -> Path:
    if is_filtered_run:
        return paths.output_dir / f"{DATE_DOT}-{paths.name}链接检查表_filtered_{STAMP}.xlsx"
    base = paths.output_dir / f"{DATE_DOT}-{paths.name}链接检查表.xlsx"
    if not base.exists():
        return base
    return paths.output_dir / f"{DATE_DOT}-{paths.name}链接检查表_{STAMP}.xlsx"


def new_exception_output_path(paths: ProjectPaths, is_filtered_run: bool = False) -> Path:
    if is_filtered_run:
        return paths.output_dir / f"{DATE_DOT}-{paths.name}异常汇总表_filtered_{STAMP}.xlsx"
    base = paths.output_dir / f"{DATE_DOT}-{paths.name}异常汇总表.xlsx"
    if not base.exists():
        return base
    return paths.output_dir / f"{DATE_DOT}-{paths.name}异常汇总表_{STAMP}.xlsx"


def row_value(row: list[Any], header: str) -> Any:
    try:
        return row[REPORT_HEADERS.index(header)]
    except ValueError:
        return ""


def unique_detail_child_count(rows: list[list[Any]]) -> int:
    children = {
        str(row_value(row, "子ASIN") or "").strip().upper()
        for row in rows
        if str(row_value(row, "子ASIN") or "").strip()
    }
    return len(children)


def is_cart_lost(value: Any) -> bool:
    text = str(value or "")
    if not text:
        return True
    if any(token in text for token in ["不可售", "无购物车", "未知", "未确认", "丢失"]):
        return True
    return not any(token in text for token in ["有购物车", "有Buy"])


def price_change_over_threshold(row: list[Any], previous: dict[str, Any], threshold: float = 0.10) -> bool:
    child = str(row_value(row, "子ASIN") or "")
    current = parse_money_value(row_value(row, "当前/Buy Box价格"))
    before = parse_money_value((previous.get(child) or {}).get("current_price"))
    if not current or not before:
        return False
    return abs(current - before) / before > threshold


def write_link_check_workbook(paths: ProjectPaths, detail_rows: list[list[Any]], previous: dict[str, Any], is_filtered_run: bool = False) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{DATE_DOT}抓取数据"
    write_header(ws, REPORT_HEADERS)
    for row in detail_rows:
        ws.append(row)
    style_ws(ws)
    embed_images(ws, paths)

    header_to_col = {cell.value: cell.column for cell in ws[1]}
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    red_font = Font(color="9C0006")
    cart_col = header_to_col.get("购物车")
    price_col = header_to_col.get("当前/Buy Box价格")
    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(REPORT_HEADERS) + 1)]
        if cart_col and is_cart_lost(row_value(row, "购物车")):
            cell = ws.cell(row=row_idx, column=cart_col)
            cell.fill = red_fill
            cell.font = red_font
        if price_col and price_change_over_threshold(row, previous):
            cell = ws.cell(row=row_idx, column=price_col)
            cell.fill = red_fill
            cell.font = red_font

    ws.auto_filter.ref = ws.dimensions
    return save_book(wb, new_output_path(paths, is_filtered_run=is_filtered_run))


def write_exception_summary_workbook(paths: ProjectPaths, exception_rows: list[dict[str, str]], is_filtered_run: bool = False) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{DATE_DOT}异常汇总"
    write_header(ws, EXCEPTION_HEADERS)
    for row in exception_rows:
        ws.append([row.get(header, "") for header in EXCEPTION_HEADERS])
    style_ws(ws)
    ws.auto_filter.ref = ws.dimensions
    return save_book(wb, new_exception_output_path(paths, is_filtered_run=is_filtered_run))


def write_total_book(paths: ProjectPaths, detail_rows: list[list[Any]], exception_rows: list[dict[str, str]]) -> Path:
    wb = ensure_book(paths.total_book)
    for name in list(wb.sheetnames):
        if name.endswith("异常汇总"):
            del wb[name]
    ws_exc = wb.create_sheet(f"{DATE}异常汇总", 0)
    write_header(ws_exc, EXCEPTION_HEADERS)
    for row in exception_rows:
        ws_exc.append([row.get(h, "") for h in EXCEPTION_HEADERS])
    style_ws(ws_exc)
    sheet_name = unique_sheet_name(wb, DATE)
    ws = wb.create_sheet(sheet_name)
    write_header(ws, REPORT_HEADERS)
    for row in detail_rows:
        ws.append(row)
    style_ws(ws)
    embed_images(ws, paths)
    reset_default_sheet(wb)
    return save_book(wb, paths.total_book)


def write_exception_book(paths: ProjectPaths, exception_rows: list[dict[str, str]]) -> Path:
    wb = ensure_book(paths.exception_book)
    sheet_name = unique_sheet_name(wb, DATE)
    ws = wb.create_sheet(sheet_name)
    write_header(ws, EXCEPTION_HEADERS)
    for row in exception_rows:
        ws.append([row.get(h, "") for h in EXCEPTION_HEADERS])
    style_ws(ws)
    reset_default_sheet(wb)
    return save_book(wb, paths.exception_book)


def write_self_check_report(paths: ProjectPaths, detail_rows: list[list[Any]], exception_rows: list[dict[str, str]], run_stats: dict[str, Any]) -> None:
    header_index = {header: idx for idx, header in enumerate(REPORT_HEADERS)}

    def value(row: list[Any], header: str) -> Any:
        idx = header_index.get(header)
        return row[idx] if idx is not None and idx < len(row) else ""

    rows = detail_rows
    report = {
        "checked_at": datetime.now().isoformat(timespec="seconds"),
        "date": DATE,
        "project": paths.name,
        "run_mode": run_stats.get("run_mode", "full"),
        "full_item_count": run_stats.get("full_item_count", 0),
        "planned_item_count": run_stats.get("planned_item_count", 0),
        "planned_unique_child_count": run_stats.get("planned_unique_child_count", 0),
        "processed_count": run_stats.get("processed_count", 0),
        "missing_processed_count": run_stats.get("missing_processed_count", 0),
        "quantity_check_ok": run_stats.get("quantity_check_ok", False),
        "filtered_run": run_stats.get("filtered_run", False),
        "stopped_early": run_stats.get("stopped_early", False),
        "ok_count": run_stats.get("ok_count", 0),
        "error_count": run_stats.get("error_count", 0),
        "captcha_count": run_stats.get("captcha_count", 0),
        "blocked_count": run_stats.get("blocked_count", 0),
        "abnormal_item_count": run_stats.get("abnormal_item_count", 0),
        "exception_issue_count": run_stats.get("exception_issue_count", 0),
        "snapshot_entry_count": run_stats.get("snapshot_entry_count", 0),
        "detail_rows": len(rows),
        "detail_unique_child_count": unique_detail_child_count(rows),
        "duplicate_child_rows": max(0, len(rows) - unique_detail_child_count(rows)),
        "exception_rows": len(exception_rows),
        "exception_issue_rows": len(exception_rows),
        "missing_current_price": sum(1 for row in rows if not value(row, "当前/Buy Box价格")),
        "suspicious_unit_price": [
            value(row, "子ASIN")
            for row in rows
            if str(value(row, "当前/Buy Box价格")) in {"$10.00", "$11.00", "$11.99", "$14.99"}
        ],
        "missing_category": sum(1 for row in rows if not value(row, "类目节点")),
        "missing_rank": sum(1 for row in rows if not value(row, "大类排名") and not value(row, "小类排名")),
        "page_without_breadcrumb_notes": sum(1 for row in rows if "页面未显示顶部breadcrumb" in str(value(row, "备注"))),
        "parser_missed_breadcrumb_notes": sum(1 for row in rows if "页面存在顶部breadcrumb标记" in str(value(row, "备注"))),
        "page_without_rank_notes": sum(1 for row in rows if "页面未显示 Best Sellers Rank" in str(value(row, "备注"))),
        "parser_missed_rank_notes": sum(1 for row in rows if "页面存在 Best Sellers Rank 标记" in str(value(row, "备注"))),
        "strike_without_percent": sum(1 for row in rows if value(row, "是否有划线") == "是" and not value(row, "划线百分比")),
        "unknown_buybox": sum(1 for row in rows if value(row, "购物车") == "未知"),
        "star_images": sum(1 for row in rows if value(row, "星级占比截图")),
        "price_images": sum(1 for row in rows if value(row, "价格截图")),
        "parent_review_split_notes": sum(1 for row in rows if "同父体评论数不一致" in str(value(row, "备注"))),
        "new_bad_review_notes": sum(1 for row in rows if "新来差评" in str(value(row, "备注"))),
    }
    out_dir = paths.cache_dir / "self_checks"
    out_dir.mkdir(parents=True, exist_ok=True)
    write_json(out_dir / f"self_check_{DATE}_{STAMP}.json", report)


def write_self_check_report(paths: ProjectPaths, detail_rows: list[list[Any]], exception_rows: list[dict[str, str]], run_stats: dict[str, Any]) -> None:
    header_index = {header: idx for idx, header in enumerate(REPORT_HEADERS)}

    def value(row: list[Any], header: str) -> Any:
        idx = header_index.get(header)
        return row[idx] if idx is not None and idx < len(row) else ""

    rows = detail_rows
    def is_invalid_link_row(row: list[Any]) -> bool:
        return "链接失效" in str(value(row, "备注"))

    def is_unavailable_buybox_row(row: list[Any]) -> bool:
        buybox = str(value(row, "购物车") or "")
        return "不可售" in buybox or "无购物车" in buybox

    zipcode_dependent_buybox = []
    for row in rows:
        match = re.search(r"购物车依赖邮编：(\d{5})", str(value(row, "备注")))
        if match:
            zipcode_dependent_buybox.append({
                "child_asin": value(row, "子ASIN"),
                "zipcode": match.group(1),
            })

    report = {
        "checked_at": datetime.now().isoformat(timespec="seconds"),
        "date": DATE,
        "project": paths.name,
        "run_mode": run_stats.get("run_mode", "full"),
        "full_item_count": run_stats.get("full_item_count", 0),
        "planned_item_count": run_stats.get("planned_item_count", 0),
        "planned_unique_child_count": run_stats.get("planned_unique_child_count", 0),
        "processed_count": run_stats.get("processed_count", 0),
        "missing_processed_count": run_stats.get("missing_processed_count", 0),
        "quantity_check_ok": run_stats.get("quantity_check_ok", False),
        "filtered_run": run_stats.get("filtered_run", False),
        "stopped_early": run_stats.get("stopped_early", False),
        "ok_count": run_stats.get("ok_count", 0),
        "error_count": run_stats.get("error_count", 0),
        "captcha_count": run_stats.get("captcha_count", 0),
        "blocked_count": run_stats.get("blocked_count", 0),
        "abnormal_item_count": run_stats.get("abnormal_item_count", 0),
        "exception_issue_count": run_stats.get("exception_issue_count", 0),
        "snapshot_entry_count": run_stats.get("snapshot_entry_count", 0),
        "detail_rows": len(rows),
        "detail_unique_child_count": unique_detail_child_count(rows),
        "duplicate_child_rows": max(0, len(rows) - unique_detail_child_count(rows)),
        "exception_rows": len(exception_rows),
        "exception_issue_rows": len(exception_rows),
        "missing_current_price": sum(
            1
            for row in rows
            if not is_invalid_link_row(row)
            and not is_unavailable_buybox_row(row)
            and not value(row, "当前/Buy Box价格")
        ),
        "unavailable_buybox_rows": sum(1 for row in rows if is_unavailable_buybox_row(row)),
        "zipcode_checked_without_buybox": [
            value(row, "子ASIN")
            for row in rows
            if is_unavailable_buybox_row(row) and "邮编检查：" in str(value(row, "备注"))
        ],
        "suspicious_unit_price": [
            value(row, "子ASIN")
            for row in rows
            if (
                str(value(row, "当前/Buy Box价格")) in {"$10.00", "$11.00", "$11.99", "$14.99"}
                and not value(row, "List Price")
                and not value(row, "Typical Price")
                and not value(row, "划线百分比")
            )
        ],
        "invalid_link_rows": sum(1 for row in rows if is_invalid_link_row(row)),
        "missing_category": sum(1 for row in rows if not is_invalid_link_row(row) and not value(row, "类目节点")),
        "missing_rank": sum(1 for row in rows if not is_invalid_link_row(row) and not value(row, "大类排名") and not value(row, "小类排名")),
        "page_without_breadcrumb_notes": sum(1 for row in rows if "页面未显示顶部breadcrumb" in str(value(row, "备注"))),
        "parser_missed_breadcrumb_notes": sum(1 for row in rows if "页面存在顶部breadcrumb标记" in str(value(row, "备注"))),
        "page_without_rank_notes": sum(1 for row in rows if "页面未显示 Best Sellers Rank" in str(value(row, "备注"))),
        "parser_missed_rank_notes": sum(1 for row in rows if "页面存在 Best Sellers Rank 标记" in str(value(row, "备注"))),
        "strike_without_percent": sum(1 for row in rows if value(row, "是否有划线") == "是" and not value(row, "划线百分比")),
        "typical_only_marked_strike": [
            value(row, "子ASIN")
            for row in rows
            if value(row, "Typical Price")
            and not value(row, "List Price")
            and (value(row, "是否有划线") == "是" or value(row, "划线百分比"))
        ],
        "unknown_buybox": sum(
            1
            for row in rows
            if not is_invalid_link_row(row)
            and (value(row, "购物车") == "未知" or "未确认" in str(value(row, "购物车") or ""))
        ),
        "overlong_promotion_text": [
            value(row, "子ASIN")
            for row in rows
            if len(str(value(row, "买赠/多买折扣") or "")) > 280
        ],
        "zipcode_dependent_buybox_count": len(zipcode_dependent_buybox),
        "zipcode_dependent_buybox": zipcode_dependent_buybox,
        "star_images": sum(1 for row in rows if value(row, "星级占比截图")),
        "price_images": sum(1 for row in rows if value(row, "价格截图")),
        "parent_review_split_notes": sum(1 for row in rows if "同父体评论数不一致" in str(value(row, "备注"))),
        "new_bad_review_notes": sum(1 for row in rows if "新来差评" in str(value(row, "备注"))),
    }
    out_dir = paths.cache_dir / "self_checks"
    out_dir.mkdir(parents=True, exist_ok=True)
    write_json(out_dir / f"self_check_{DATE}_{STAMP}.json", report)


def should_update_snapshot(results: dict[str, Any], total: int) -> bool:
    if not total:
        return False
    valid = sum(1 for item in results.values() if item.get("status") == "OK" and item.get("title") and item.get("reviews"))
    captcha = sum(1 for item in results.values() if item.get("captcha"))
    return valid >= max(1, total // 2) and captcha / total < 0.4


def persist_run_progress(paths: ProjectPaths, progress: dict[str, Any]) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    progress["checked_at"] = now
    progress["updated_at"] = now
    write_json(paths.cache_dir / "run_progress.json", progress)


class MemoryStatus(Structure):
    _fields_ = [
        ("length", c_ulong),
        ("memory_load", c_ulong),
        ("total_physical", c_ulonglong),
        ("available_physical", c_ulonglong),
        ("total_page_file", c_ulonglong),
        ("available_page_file", c_ulonglong),
        ("total_virtual", c_ulonglong),
        ("available_virtual", c_ulonglong),
        ("available_extended_virtual", c_ulonglong),
    ]


def available_memory_gb() -> float | None:
    try:
        status = MemoryStatus()
        status.length = sizeof(MemoryStatus)
        if windll.kernel32.GlobalMemoryStatusEx(byref(status)):
            return status.available_physical / (1024 ** 3)
    except Exception:
        pass
    return None


def determine_batch_workers(batch_count: int) -> tuple[int, dict[str, Any]]:
    if batch_count <= 0:
        return 0, {"mode": "none", "cpu_count": cpu_count() or 1, "available_memory_gb": available_memory_gb()}
    processors = cpu_count() or 1
    memory_gb = available_memory_gb()
    cpu_limit = max(1, processors // 4)
    memory_limit = MAX_BATCH_WORKERS if memory_gb is None else max(1, int(max(0.0, memory_gb - 1.5) // 1.5))
    automatic = max(1, min(MAX_BATCH_WORKERS, cpu_limit, memory_limit, batch_count))
    selected = automatic if ASIN_BATCH_WORKERS == 0 else max(1, min(ASIN_BATCH_WORKERS, MAX_BATCH_WORKERS, batch_count, memory_limit))
    return selected, {
        "mode": "auto" if ASIN_BATCH_WORKERS == 0 else "configured",
        "cpu_count": processors,
        "available_memory_gb": round(memory_gb, 2) if memory_gb is not None else None,
        "cpu_limit": cpu_limit,
        "memory_limit": memory_limit,
        "hard_limit": MAX_BATCH_WORKERS,
    }


def batch_run_key(paths: ProjectPaths, items: list[dict[str, str]]) -> str:
    stat = paths.input_file.stat()
    payload = {
        "project": paths.name,
        "input_file": str(paths.input_file),
        "input_size": stat.st_size,
        "input_mtime_ns": stat.st_mtime_ns,
        "items": items,
        "postal_codes": POSTAL_CODES,
        "batch_size": ASIN_BATCH_SIZE,
    }
    digest = hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()
    return f"{DATE}_{digest[:12]}"


def create_batches(items: list[dict[str, str]], run_dir: Path) -> list[dict[str, Any]]:
    batches = []
    ranges = []
    if items:
        ranges.append((0, 1))
        ranges.extend((offset, min(offset + ASIN_BATCH_SIZE, len(items))) for offset in range(1, len(items), ASIN_BATCH_SIZE))
    for offset, end in ranges:
        number = len(batches) + 1
        entries = [{"index": index, "item": items[index]} for index in range(offset, end)]
        batch_dir = run_dir / f"batch_{number:04d}"
        batches.append({
            "batch_number": number,
            "batch_id": f"batch_{number:04d}",
            "batch_dir": batch_dir,
            "entries": entries,
            "planned_count": len(entries),
            "result_file": batch_dir / "results.json",
            "progress_file": batch_dir / "progress.json",
        })
    return batches


def completed_batch_records(batch: dict[str, Any]) -> list[dict[str, Any]] | None:
    payload = saved_batch_payload(batch)
    if not payload:
        return None
    records = payload.get("records", [])
    if payload.get("status") == "completed" and len(records) == batch["planned_count"]:
        return records
    return None


def saved_batch_payload(batch: dict[str, Any]) -> dict[str, Any] | None:
    result_file = batch["result_file"]
    if not ASIN_BATCH_RESUME or not result_file.exists():
        return None
    try:
        payload = json.loads(result_file.read_text(encoding="utf-8"))
    except Exception:
        return None
    expected_indexes = {entry["index"] for entry in batch["entries"]}
    records = payload.get("records", [])
    record_indexes = [record.get("index") for record in records]
    if len(record_indexes) != len(set(record_indexes)) or any(index not in expected_indexes for index in record_indexes):
        return None
    return payload


def capture_failure_reason(record: dict[str, Any] | None) -> str:
    if not record:
        return "no capture result"
    current = record.get("current", {})
    if current.get("not_found"):
        return ""
    if current.get("captcha"):
        return "CAPTCHA detected"
    if current.get("blocked"):
        return "Amazon page blocked"
    if current.get("status") == "ERROR":
        return current.get("error") or "capture status ERROR"
    if not clean_text(current.get("title")):
        return "product title missing; page capture is not usable"
    return ""


def run_batch(paths: ProjectPaths, batch: dict[str, Any], progress_callback) -> dict[str, Any]:
    started_at = datetime.now().isoformat(timespec="seconds")
    saved_payload = saved_batch_payload(batch) or {}
    records: list[dict[str, Any]] = list(saved_payload.get("records", []))
    completed_indexes = {record["index"] for record in records}
    resumed_seconds = sum(float(record.get("current", {}).get("capture_seconds", 0) or 0) for record in records)
    progress = {
        "batch_id": batch["batch_id"],
        "batch_number": batch["batch_number"],
        "status": "running",
        "started_at": started_at,
        "updated_at": started_at,
        "planned_count": batch["planned_count"],
        "processed_count": len(records),
        "remaining_count": batch["planned_count"] - len(records),
        "ok_count": sum(record["current"].get("status") != "ERROR" and not record["current"].get("captcha") for record in records),
        "error_count": sum(record["current"].get("status") == "ERROR" for record in records),
        "captcha_count": sum(bool(record["current"].get("captcha")) for record in records),
        "blocked_count": sum(bool(record["current"].get("blocked")) for record in records),
        "total_capture_seconds": round(resumed_seconds, 2),
        "average_capture_seconds": round(resumed_seconds / len(records), 2) if records else 0.0,
        "max_capture_seconds": max((float(record.get("current", {}).get("capture_seconds", 0) or 0) for record in records), default=0.0),
        "resumed_record_count": len(records),
        "last_child_asin": item_value(records[-1]["item"], "子ASIN") if records else "",
        "error": "",
    }
    write_json(batch["progress_file"], progress)
    progress_callback(batch["batch_id"], progress)
    console_log(
        f"{paths.name}/{batch['batch_id']}",
        f"started planned={batch['planned_count']} resumed={len(records)} remaining={progress['remaining_count']}",
    )
    consecutive_blocks = 0
    batch_paths = replace(paths, screenshot_dir=batch["batch_dir"] / "screenshots")
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(locale="en-US", viewport={"width": 1440, "height": 1200})
            try:
                for entry in batch["entries"]:
                    if entry["index"] in completed_indexes:
                        continue
                    item = entry["item"]
                    child = item_value(item, "子ASIN")
                    page = context.new_page()
                    page.set_default_timeout(25000)
                    page.set_default_navigation_timeout(65000)
                    capture_started = perf_counter()
                    try:
                        current = extract_with_zip_fallback(page, item, batch_paths)
                    except PlaywrightTimeoutError as exc:
                        current = {"status": "ERROR", "captcha": False, "title": "", "reviews": "", "rank": "", "error": str(exc)}
                    except Exception as exc:
                        current = {"status": "ERROR", "captcha": False, "title": "", "reviews": "", "rank": "", "error": str(exc)}
                    finally:
                        try:
                            page.close()
                        except Exception:
                            pass
                    capture_seconds = round(perf_counter() - capture_started, 2)
                    current["capture_seconds"] = capture_seconds
                    records.append({"index": entry["index"], "item": item, "current": current})
                    consecutive_blocks = consecutive_blocks + 1 if current.get("captcha") or current.get("blocked") else 0
                    if current.get("status") == "ERROR":
                        progress["error_count"] += 1
                    elif current.get("captcha"):
                        progress["captcha_count"] += 1
                    else:
                        progress["ok_count"] += 1
                    if current.get("blocked"):
                        progress["blocked_count"] += 1
                    progress["total_capture_seconds"] = round(progress["total_capture_seconds"] + capture_seconds, 2)
                    progress["average_capture_seconds"] = round(progress["total_capture_seconds"] / len(records), 2)
                    progress["max_capture_seconds"] = max(progress["max_capture_seconds"], capture_seconds)
                    progress["processed_count"] = len(records)
                    progress["remaining_count"] = batch["planned_count"] - len(records)
                    progress["last_child_asin"] = child
                    progress["last_status"] = current.get("status", "")
                    progress["updated_at"] = datetime.now().isoformat(timespec="seconds")
                    write_json(batch["progress_file"], progress)
                    progress_callback(batch["batch_id"], progress)
                    console_log(
                        f"{paths.name}/{batch['batch_id']}",
                        f"progress={len(records)}/{batch['planned_count']} child={child} status={current.get('status', '')} "
                        f"seconds={capture_seconds:.2f} avg_seconds={progress['average_capture_seconds']:.2f} "
                        f"errors={progress['error_count']} captcha={progress['captcha_count']} blocked={progress['blocked_count']}",
                    )
                    if consecutive_blocks >= MAX_CONSECUTIVE_BLOCKS:
                        progress["status"] = "stopped_early"
                        progress["error"] = f"{consecutive_blocks} consecutive blocked pages"
                        break
            finally:
                context.close()
                browser.close()
        if progress["status"] == "running":
            progress["status"] = "completed"
    except Exception as exc:
        progress["status"] = "failed"
        progress["error"] = str(exc)
    progress["finished_at"] = datetime.now().isoformat(timespec="seconds")
    progress["updated_at"] = progress["finished_at"]
    write_json(batch["progress_file"], progress)
    result = {**progress, "records": records}
    write_json(batch["result_file"], result)
    progress_callback(batch["batch_id"], progress)
    console_log(
        f"{paths.name}/{batch['batch_id']}",
        f"finished status={progress['status']} processed={progress['processed_count']}/{batch['planned_count']} error={progress.get('error', '')}",
        "ERROR" if progress["status"] in {"failed", "stopped_early"} else "INFO",
    )
    return result


def build_report_data(paths: ProjectPaths, records: list[dict[str, Any]], previous: dict[str, Any]) -> tuple[list[list[Any]], list[dict[str, str]], dict[str, Any], list[dict[str, str]]]:
    detail_rows: list[list[Any]] = []
    exception_rows: list[dict[str, str]] = []
    snapshot: dict[str, Any] = {}
    zipcode_dependent_buybox: list[dict[str, str]] = []
    parent_review_cache: dict[str, dict[str, Any]] = {}
    for record in sorted(records, key=lambda value: value["index"]):
        item = record["item"]
        current = record["current"]
        child = item_value(item, "子ASIN")
        parent = item_value(item, "父ASIN")
        prior_parent = parent_review_cache.get(parent)
        if prior_parent:
            first_seen_child = prior_parent.get("first_seen_child") or prior_parent.get("child") or ""
            first_seen_reviews = prior_parent.get("first_seen_reviews", "")
            current["parent_first_child_ref"] = f"首个子体{first_seen_child}={first_seen_reviews}"
            cached_reviews = parse_int(prior_parent.get("reviews"))
            current_reviews = parse_int(current.get("reviews"))
            if not cached_reviews and current_reviews:
                prior_parent["reviews"] = current.get("reviews")
                prior_parent["child"] = child
                current["parent_first_valid_child_ref"] = f"首个有效评论子体{child}={current.get('reviews')}"
            elif cached_reviews and cached_reviews == current_reviews:
                current["star_distribution_screenshot"] = ""
            elif cached_reviews:
                current["parent_review_split"] = f"同父体评论数不一致：首个有效子体{prior_parent.get('child')}={prior_parent.get('reviews')}，当前{child}={current.get('reviews')}"
        else:
            parent_review_cache[parent] = {
                "reviews": current.get("reviews"),
                "child": child,
                "first_seen_child": child,
                "first_seen_reviews": current.get("reviews"),
            }
        issues, notes = compare(item, current, previous)
        extra_note_parts = []
        if current.get("parent_first_valid_child_ref"):
            extra_note_parts.append(current["parent_first_valid_child_ref"])
        if current.get("parent_first_child_ref") and (current.get("parent_review_split") or current.get("status") == "ERROR"):
            extra_note_parts.append(current["parent_first_child_ref"])
        exception_rows.extend(issues)
        previous_rank = (previous.get(child) or {}).get("rank", "")
        big_rank, _ = rank_cell(current.get("rank", ""), previous_rank, 0)
        small_rank, _ = rank_cell(current.get("rank", ""), previous_rank, 1)
        zip_note = ""
        if current.get("delivery_zip_checked"):
            zip_note = f"邮编检查：{current.get('delivery_zip_checked')}"
            if current.get("buybox_zip_dependent") and current.get("buybox_zip_success"):
                zip_note += f"；购物车依赖邮编：{current.get('buybox_zip_success')}"
        detail_rows.append([
            DATE, parent, item_value(item, "子SKU"), child, "是" if issues else "否",
            current.get("buybox", "未知"), copy_text(current), current.get("aplus_visible", "否"),
            current.get("rating", ""), current.get("reviews", ""),
            f"{current.get('review_delta', ''):+d}" if isinstance(current.get("review_delta"), int) else "",
            current.get("star_distribution_screenshot", ""),
            "是" if any(row["问题模块"] == "新增差评" for row in issues) else "否",
            current.get("other_sellers", "无明显跟卖"), current.get("category", ""), big_rank, small_rank,
            current.get("list_price", ""), current.get("typical_price", ""), current.get("has_strike", ""),
            current.get("current_price", ""), current.get("discount", ""), current.get("prime", ""),
            current.get("coupon", ""), current.get("multi_buy", ""), current.get("price_screenshot", ""),
            readable_issue_notes(
                issues,
                [notes] if notes else [],
                [*extra_note_parts, zip_note, current.get("error", ""), item_value(item, "备注")],
            ),
            "否" if issues else "无需处理",
        ])
        snapshot[child] = {
            "child": child, "parent": parent, "title": current.get("title", ""), "reviews": current.get("reviews", ""),
            "rating": current.get("rating", ""), "rank": current.get("rank", ""), "current_price": current.get("current_price", ""),
            "buybox": current.get("buybox", ""), "aplus_visible": current.get("aplus_visible", ""),
            "negative_estimate": current.get("negative_estimate", 0), "checked_at": datetime.now().isoformat(timespec="seconds"),
            "status": current.get("status", ""), "captcha": bool(current.get("captcha")), "blocked": bool(current.get("blocked")),
            "error": current.get("error", ""), "delivery_zip_checked": current.get("delivery_zip_checked", ""),
            "buybox_zip_dependent": bool(current.get("buybox_zip_dependent")), "buybox_zip_success": current.get("buybox_zip_success", ""),
        }
        if current.get("buybox_zip_dependent") and current.get("buybox_zip_success") and not any(entry["child_asin"] == child for entry in zipcode_dependent_buybox):
            zipcode_dependent_buybox.append({"child_asin": child, "zipcode": current.get("buybox_zip_success"), "checked_path": current.get("delivery_zip_checked", "")})
    return detail_rows, exception_rows, snapshot, zipcode_dependent_buybox


def run_project_legacy(paths: ProjectPaths) -> dict[str, Any]:
    items = read_items(paths.input_file)
    full_item_count = len(items)
    is_filtered_run = bool(ASIN_PARENT_FILTER or ASIN_CHILD_FILTER or ASIN_LIMIT > 0)
    if ASIN_PARENT_FILTER:
        items = [item for item in items if item.get("父ASIN") == ASIN_PARENT_FILTER]
    if ASIN_CHILD_FILTER:
        items = [item for item in items if item.get("子ASIN") in ASIN_CHILD_FILTER]
    if ASIN_LIMIT > 0:
        items = items[:ASIN_LIMIT]
    previous = load_previous(paths)
    planned_item_count = len(items)
    planned_unique_child_count = len({item_value(item, "子ASIN").strip().upper() for item in items if item_value(item, "子ASIN").strip()})
    detail_rows: list[list[Any]] = []
    exception_rows: list[dict[str, str]] = []
    snapshot: dict[str, Any] = {}
    parent_review_cache: dict[str, dict[str, Any]] = {}
    started_at = datetime.now().isoformat(timespec="seconds")
    run_progress = {
        "started_at": started_at,
        "checked_at": started_at,
        "date": DATE,
        "project": paths.name,
        "script_path": str(Path(__file__)),
        "run_mode": run_mode_label(is_filtered_run),
        "full_item_count": full_item_count,
        "planned_item_count": planned_item_count,
        "planned_unique_child_count": planned_unique_child_count,
        "processed_count": 0,
        "remaining_count": planned_item_count,
        "ok_count": 0,
        "error_count": 0,
        "captcha_count": 0,
        "blocked_count": 0,
        "abnormal_item_count": 0,
        "exception_issue_count": 0,
        "zipcode_dependent_buybox": [],
        "filtered_run": is_filtered_run,
        "completed": False,
        "stopped_early": False,
    }
    persist_run_progress(paths, run_progress)
    print(f"[{paths.name}] 计划抓取 {planned_item_count}/{full_item_count} 条")

    if sync_playwright is None:
        raise RuntimeError("Playwright 未安装。请先执行: python -m pip install playwright && python -m playwright install chromium")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(locale="en-US", viewport={"width": 1440, "height": 1200})
        consecutive_blocks = 0
        for idx, item in enumerate(items, start=1):
            child = item_value(item, "子ASIN")
            parent = item_value(item, "父ASIN")
            page = context.new_page()
            page.set_default_timeout(25000)
            page.set_default_navigation_timeout(65000)
            try:
                current = extract_with_zip_fallback(page, item, paths)
            except PlaywrightTimeoutError as exc:
                current = {"status": "ERROR", "captcha": False, "title": "", "reviews": "", "rank": "", "error": str(exc)}
            except Exception as exc:
                current = {"status": "ERROR", "captcha": False, "title": "", "reviews": "", "rank": "", "error": str(exc)}
            finally:
                try:
                    page.close()
                except Exception:
                    pass

            if current.get("captcha") or current.get("blocked"):
                consecutive_blocks += 1
            else:
                consecutive_blocks = 0

            prior_parent = parent_review_cache.get(parent)
            if prior_parent:
                first_seen_child = prior_parent.get("first_seen_child") or prior_parent.get("child") or ""
                first_seen_reviews = prior_parent.get("first_seen_reviews", "")
                current["parent_first_child_ref"] = f"首个子体{first_seen_child}={first_seen_reviews}"
                cached_reviews = parse_int(prior_parent.get("reviews"))
                current_reviews = parse_int(current.get("reviews"))
                if not cached_reviews and current_reviews:
                    prior_parent["reviews"] = current.get("reviews")
                    prior_parent["child"] = child
                    current["parent_first_valid_child_ref"] = f"首个有效评论子体{child}={current.get('reviews')}"
                elif cached_reviews and cached_reviews == current_reviews:
                    current["star_distribution_screenshot"] = ""
                elif cached_reviews:
                    current["parent_review_split"] = f"同父体评论数不一致：首个有效子体{prior_parent.get('child')}={prior_parent.get('reviews')}，当前{child}={current.get('reviews')}"
            else:
                parent_review_cache[parent] = {
                    "reviews": current.get("reviews"),
                    "child": child,
                    "first_seen_child": child,
                    "first_seen_reviews": current.get("reviews"),
                }

            issues, notes = compare(item, current, previous)
            note_parts = [notes] if notes else []
            if current.get("parent_first_valid_child_ref"):
                note_parts.append(current["parent_first_valid_child_ref"])
            if current.get("parent_first_child_ref") and (current.get("parent_review_split") or current.get("status") == "ERROR"):
                note_parts.append(current["parent_first_child_ref"])
            exception_rows.extend(issues)
            issue_notes = issue_summary_text(issues)
            previous_rank = (previous.get(child) or {}).get("rank", "")
            big_rank, _ = rank_cell(current.get("rank", ""), previous_rank, 0)
            small_rank, _ = rank_cell(current.get("rank", ""), previous_rank, 1)
            is_bad_review_new = "是" if any(row["问题模块"] == "新增差评" for row in issues) else "否"
            price_shot = current.get("price_screenshot", "")
            review_shot = current.get("star_distribution_screenshot", "")
            zip_note = ""
            if current.get("delivery_zip_checked"):
                zip_note = f"邮编检查：{current.get('delivery_zip_checked')}"
                if current.get("buybox_zip_dependent") and current.get("buybox_zip_success"):
                    zip_note += f"；购物车依赖邮编：{current.get('buybox_zip_success')}"
            detail_rows.append([
                DATE,
                parent,
                item_value(item, "子SKU"),
                child,
                "是" if issues else "否",
                current.get("buybox", "未知"),
                copy_text(current),
                current.get("aplus_visible", "否"),
                current.get("rating", ""),
                current.get("reviews", ""),
                f"{current.get('review_delta', ''):+d}" if isinstance(current.get("review_delta"), int) else "",
                review_shot,
                is_bad_review_new,
                current.get("other_sellers", "无明显跟卖"),
                current.get("category", ""),
                big_rank,
                small_rank,
                current.get("list_price", ""),
                current.get("typical_price", ""),
                current.get("has_strike", ""),
                current.get("current_price", ""),
                current.get("discount", ""),
                current.get("prime", ""),
                current.get("coupon", ""),
                current.get("multi_buy", ""),
                price_shot,
                "；".join([v for v in [issue_notes, "；".join(note_parts), zip_note, current.get("error", ""), item_value(item, "备注")] if v]),
                "否" if issues else "无需处理",
            ])
            snapshot[child] = {
                "child": child,
                "parent": parent,
                "title": current.get("title", ""),
                "reviews": current.get("reviews", ""),
                "rating": current.get("rating", ""),
                "rank": current.get("rank", ""),
                "current_price": current.get("current_price", ""),
                "buybox": current.get("buybox", ""),
                "aplus_visible": current.get("aplus_visible", ""),
                "negative_estimate": current.get("negative_estimate", 0),
                "checked_at": datetime.now().isoformat(timespec="seconds"),
                "status": current.get("status", ""),
                "captcha": bool(current.get("captcha")),
                "blocked": bool(current.get("blocked")),
                "error": current.get("error", ""),
                "delivery_zip_checked": current.get("delivery_zip_checked", ""),
                "buybox_zip_dependent": bool(current.get("buybox_zip_dependent")),
                "buybox_zip_success": current.get("buybox_zip_success", ""),
            }
            if current.get("buybox_zip_dependent") and current.get("buybox_zip_success"):
                if not any(entry.get("child_asin") == child for entry in run_progress["zipcode_dependent_buybox"]):
                    run_progress["zipcode_dependent_buybox"].append({
                        "child_asin": child,
                        "zipcode": current.get("buybox_zip_success"),
                        "checked_path": current.get("delivery_zip_checked", ""),
                    })
            if current.get("status") == "ERROR":
                run_progress["error_count"] += 1
            elif current.get("captcha"):
                run_progress["captcha_count"] += 1
            else:
                run_progress["ok_count"] += 1
            if current.get("blocked"):
                run_progress["blocked_count"] += 1
            if issues:
                run_progress["abnormal_item_count"] += 1
            run_progress["exception_issue_count"] += len(issues)
            run_progress["processed_count"] = idx
            run_progress["remaining_count"] = max(0, planned_item_count - idx)
            run_progress["last_parent_asin"] = parent
            run_progress["last_child_asin"] = child
            run_progress["last_status"] = current.get("status", "")
            run_progress["last_issue_count"] = len(issues)
            persist_run_progress(paths, run_progress)
            print(f"[{paths.name}] progress {idx}/{planned_item_count} child={child} status={current.get('status', '')} issues={len(issues)}")
            if consecutive_blocks >= MAX_CONSECUTIVE_BLOCKS:
                run_progress["stopped_early"] = True
                print(f"WARNING: stopped early after {consecutive_blocks} consecutive blocked pages; writing partial workbook.", file=sys.stderr)
                break
        context.close()
        browser.close()

    detail_book_path = write_link_check_workbook(paths, detail_rows, previous, is_filtered_run=is_filtered_run)
    exception_book_path = write_exception_summary_workbook(paths, exception_rows, is_filtered_run=is_filtered_run)
    run_progress["completed"] = True
    run_progress["finished_at"] = datetime.now().isoformat(timespec="seconds")
    run_progress["detail_row_count"] = len(detail_rows)
    run_progress["detail_unique_child_count"] = unique_detail_child_count(detail_rows)
    run_progress["snapshot_entry_count"] = len(snapshot)
    run_progress["missing_processed_count"] = max(0, planned_item_count - len(detail_rows))
    run_progress["duplicate_child_rows"] = max(0, len(detail_rows) - run_progress["detail_unique_child_count"])
    run_progress["quantity_check_ok"] = len(detail_rows) == planned_item_count and len(snapshot) == run_progress["detail_unique_child_count"]
    persist_run_progress(paths, run_progress)
    write_json(paths.cache_dir / f"run_progress_{DATE}_{STAMP}.json", run_progress)
    write_self_check_report(paths, detail_rows, exception_rows, run_progress)
    paths.cache_dir.mkdir(parents=True, exist_ok=True)
    write_json(paths.cache_dir / f"raw_results_{DATE}_{STAMP}.json", snapshot)
    snapshot_quality = "low_quality"
    if is_filtered_run:
        write_json(paths.cache_dir / f"snapshot_filtered_{DATE}_{STAMP}.json", snapshot)
        snapshot_quality = "filtered"
    elif len(items) == full_item_count and should_update_snapshot(snapshot, len(items)):
        write_json(paths.snapshot_file, snapshot)
        write_json(paths.cache_dir / f"snapshot_{DATE}_{STAMP}.json", snapshot)
        snapshot_quality = "snapshot"
    else:
        write_json(paths.cache_dir / f"snapshot_low_quality_{DATE}_{STAMP}.json", snapshot)
    return {
        "project": paths.name,
        "input_file": str(paths.input_file),
        "script_path": str(Path(__file__)),
        "run_mode": run_mode_label(is_filtered_run),
        "full_item_count": full_item_count,
        "planned_item_count": planned_item_count,
        "planned_unique_child_count": planned_unique_child_count,
        "processed_count": len(detail_rows),
        "missing_processed_count": run_progress["missing_processed_count"],
        "ok_count": run_progress["ok_count"],
        "error_count": run_progress["error_count"],
        "captcha_count": run_progress["captcha_count"],
        "blocked_count": run_progress["blocked_count"],
        "abnormal_item_count": run_progress["abnormal_item_count"],
        "exception_issue_count": run_progress["exception_issue_count"],
        "zipcode_dependent_buybox": run_progress["zipcode_dependent_buybox"],
        "quantity_check_ok": run_progress["quantity_check_ok"],
        "stopped_early": run_progress["stopped_early"],
        "filtered_run": is_filtered_run,
        "detail_workbook": str(detail_book_path),
        "exception_workbook": str(exception_book_path),
        "snapshot_quality": snapshot_quality,
        "snapshot_entries": len(snapshot),
    }


def run_project(paths: ProjectPaths) -> dict[str, Any]:
    items = read_items(paths.input_file)
    full_item_count = len(items)
    is_filtered_run = bool(ASIN_PARENT_FILTER or ASIN_CHILD_FILTER or ASIN_LIMIT > 0)
    if ASIN_PARENT_FILTER:
        items = [item for item in items if item.get("父ASIN") == ASIN_PARENT_FILTER]
    if ASIN_CHILD_FILTER:
        items = [item for item in items if item.get("子ASIN") in ASIN_CHILD_FILTER]
    if ASIN_LIMIT > 0:
        items = items[:ASIN_LIMIT]
    if sync_playwright is None:
        raise RuntimeError("Playwright 未安装。请先执行 python -m pip install playwright && python -m playwright install chromium")

    previous = load_previous(paths)
    planned_item_count = len(items)
    planned_unique_child_count = len({item_value(item, "子ASIN").strip().upper() for item in items if item_value(item, "子ASIN").strip()})
    run_key = batch_run_key(paths, items)
    if not ASIN_BATCH_RESUME:
        run_key = f"{run_key}_{STAMP}"
    run_dir = paths.cache_dir / "batch_runs" / run_key
    batches = create_batches(items, run_dir)
    worker_count, resource_decision = determine_batch_workers(len(batches))
    started_at = datetime.now().isoformat(timespec="seconds")
    aggregate_file = run_dir / "aggregate_progress.json"
    latest_aggregate_file = paths.cache_dir / "batch_progress.json"
    aggregate_lock = Lock()
    aggregate = {
        "run_key": run_key,
        "date": DATE,
        "project": paths.name,
        "script_path": str(Path(__file__)),
        "run_mode": run_mode_label(is_filtered_run),
        "started_at": started_at,
        "updated_at": started_at,
        "status": "running",
        "batch_size": ASIN_BATCH_SIZE,
        "batch_count": len(batches),
        "worker_count": worker_count,
        "resource_decision": resource_decision,
        "resume_enabled": ASIN_BATCH_RESUME,
        "postal_codes": POSTAL_CODES,
        "planned_item_count": planned_item_count,
        "processed_count": 0,
        "remaining_count": planned_item_count,
        "completed_batch_count": 0,
        "failed_batch_count": 0,
        "active_batch_count": 0,
        "batches": {
            batch["batch_id"]: {
                "batch_number": batch["batch_number"],
                "planned_count": batch["planned_count"],
                "processed_count": 0,
                "remaining_count": batch["planned_count"],
                "status": "pending",
                "progress_file": str(batch["progress_file"]),
                "result_file": str(batch["result_file"]),
            }
            for batch in batches
        },
    }

    def persist_aggregate() -> None:
        states = list(aggregate["batches"].values())
        aggregate["processed_count"] = sum(state.get("processed_count", 0) for state in states)
        aggregate["remaining_count"] = max(0, planned_item_count - aggregate["processed_count"])
        aggregate["completed_batch_count"] = sum(state.get("status") == "completed" for state in states)
        aggregate["failed_batch_count"] = sum(state.get("status") in {"failed", "stopped_early"} for state in states)
        aggregate["active_batch_count"] = sum(state.get("status") == "running" for state in states)
        aggregate["total_capture_seconds"] = round(sum(state.get("total_capture_seconds", 0) for state in states), 2)
        aggregate["average_capture_seconds"] = round(
            aggregate["total_capture_seconds"] / aggregate["processed_count"], 2
        ) if aggregate["processed_count"] else 0.0
        aggregate["max_capture_seconds"] = max((state.get("max_capture_seconds", 0) for state in states), default=0.0)
        aggregate["updated_at"] = datetime.now().isoformat(timespec="seconds")
        write_json(aggregate_file, aggregate)
        write_json(latest_aggregate_file, aggregate)

    def batch_progress_callback(batch_id: str, progress: dict[str, Any]) -> None:
        with aggregate_lock:
            state = aggregate["batches"][batch_id]
            for key in [
                "status", "started_at", "finished_at", "updated_at", "processed_count", "remaining_count",
                "ok_count", "error_count", "captcha_count", "blocked_count", "last_child_asin", "last_status", "error",
                "total_capture_seconds", "average_capture_seconds", "max_capture_seconds",
                "resumed_record_count",
            ]:
                if key in progress:
                    state[key] = progress[key]
            persist_aggregate()

    console_log(
        paths.name,
        f"project started planned={planned_item_count}/{full_item_count} batches={len(batches)} "
        f"batch_size={ASIN_BATCH_SIZE} workers={worker_count} resume={ASIN_BATCH_RESUME} postal_codes={','.join(POSTAL_CODES)}",
    )
    all_records: list[dict[str, Any]] = []
    pending_batches = []
    for batch in batches:
        records = completed_batch_records(batch)
        if records is None:
            pending_batches.append(batch)
            continue
        all_records.extend(records)
        aggregate["batches"][batch["batch_id"]].update({
            "status": "completed",
            "processed_count": len(records),
            "remaining_count": 0,
            "resumed": True,
        })
    persist_aggregate()

    if not batches:
        aggregate["status"] = "aborted"
        aggregate["failure_stage"] = "project_start_gate"
        aggregate["error"] = "project contains no enabled ASIN"
        persist_aggregate()
        raise RuntimeError(f"[{paths.name}] project contains no enabled ASIN; stopping before next project")

    startup_batch = batches[0]
    if startup_batch in pending_batches:
        startup_result = run_batch(paths, startup_batch, batch_progress_callback)
        all_records.extend(startup_result.get("records", []))
        pending_batches.remove(startup_batch)
    startup_record = next((record for record in all_records if record.get("index") == 0), None)
    startup_failure = capture_failure_reason(startup_record)
    if startup_failure:
        aggregate["status"] = "aborted"
        aggregate["failure_stage"] = "project_start_gate"
        aggregate["error"] = startup_failure
        aggregate["finished_at"] = datetime.now().isoformat(timespec="seconds")
        persist_aggregate()
        console_log(paths.name, f"startup gate failed: {startup_failure}", "ERROR")
        raise RuntimeError(f"[{paths.name}] first ASIN capture failed: {startup_failure}; stopping before batch fan-out")
    console_log(paths.name, "startup gate passed; remaining batches may start")

    if pending_batches:
        with ThreadPoolExecutor(max_workers=worker_count, thread_name_prefix=f"asin-{safe_name(paths.name)}") as executor:
            futures = {executor.submit(run_batch, paths, batch, batch_progress_callback): batch for batch in pending_batches}
            for future in as_completed(futures):
                batch = futures[future]
                try:
                    result = future.result()
                except Exception as exc:
                    result = {"status": "failed", "records": [], "error": str(exc)}
                    batch_progress_callback(batch["batch_id"], result)
                all_records.extend(result.get("records", []))

    records_by_index = {record["index"]: record for record in all_records}
    ordered_records = [records_by_index[index] for index in sorted(records_by_index)]
    detail_rows, exception_rows, snapshot, zipcode_dependent_buybox = build_report_data(paths, ordered_records, previous)
    detail_book_path = write_link_check_workbook(paths, detail_rows, previous, is_filtered_run=is_filtered_run)
    exception_book_path = write_exception_summary_workbook(paths, exception_rows, is_filtered_run=is_filtered_run)

    statuses = [state["status"] for state in aggregate["batches"].values()]
    completed = len(ordered_records) == planned_item_count and all(status == "completed" for status in statuses)
    capture_error_count = sum(record["current"].get("status") == "ERROR" for record in ordered_records)
    link_invalid_count = sum(bool(record["current"].get("not_found")) for record in ordered_records)
    capture_failure_count = sum(
        record["current"].get("status") == "ERROR" and not record["current"].get("not_found")
        for record in ordered_records
    )
    capture_captcha_count = sum(bool(record["current"].get("captcha")) for record in ordered_records)
    capture_blocked_count = sum(bool(record["current"].get("blocked")) for record in ordered_records)
    project_success = completed and capture_failure_count == 0 and capture_captcha_count == 0 and capture_blocked_count == 0
    aggregate["status"] = "completed" if project_success else "partial"
    aggregate["project_success"] = project_success
    aggregate["finished_at"] = datetime.now().isoformat(timespec="seconds")
    aggregate["detail_row_count"] = len(detail_rows)
    aggregate["abnormal_item_count"] = sum(1 for row in detail_rows if row[REPORT_HEADERS.index("是否异常")] == "是")
    aggregate["exception_issue_count"] = len(exception_rows)
    aggregate["zipcode_dependent_buybox"] = zipcode_dependent_buybox
    persist_aggregate()
    console_log(
        paths.name,
        f"project collection finished success={project_success} processed={len(detail_rows)}/{planned_item_count} "
        f"capture_failures={capture_failure_count} invalid_links={link_invalid_count} "
        f"captcha={capture_captcha_count} blocked={capture_blocked_count} "
        f"abnormal_items={aggregate['abnormal_item_count']} issues={len(exception_rows)}",
        "INFO" if project_success else "ERROR",
    )

    run_progress = {
        "started_at": started_at,
        "finished_at": aggregate["finished_at"],
        "date": DATE,
        "project": paths.name,
        "script_path": str(Path(__file__)),
        "run_mode": run_mode_label(is_filtered_run),
        "full_item_count": full_item_count,
        "planned_item_count": planned_item_count,
        "planned_unique_child_count": planned_unique_child_count,
        "processed_count": len(detail_rows),
        "remaining_count": max(0, planned_item_count - len(detail_rows)),
        "ok_count": sum(record["current"].get("status") != "ERROR" and not record["current"].get("captcha") for record in ordered_records),
        "error_count": capture_error_count,
        "capture_failure_count": capture_failure_count,
        "link_invalid_count": link_invalid_count,
        "captcha_count": capture_captcha_count,
        "blocked_count": capture_blocked_count,
        "abnormal_item_count": aggregate["abnormal_item_count"],
        "exception_issue_count": len(exception_rows),
        "zipcode_dependent_buybox": zipcode_dependent_buybox,
        "filtered_run": is_filtered_run,
        "completed": completed,
        "project_success": project_success,
        "stopped_early": not completed,
        "batch_run_key": run_key,
        "batch_size": ASIN_BATCH_SIZE,
        "batch_count": len(batches),
        "batch_worker_count": worker_count,
        "batch_progress_file": str(latest_aggregate_file),
        "batch_run_progress_file": str(aggregate_file),
        "average_capture_seconds": aggregate["average_capture_seconds"],
        "max_capture_seconds": aggregate["max_capture_seconds"],
        "detail_row_count": len(detail_rows),
        "detail_unique_child_count": unique_detail_child_count(detail_rows),
        "snapshot_entry_count": len(snapshot),
        "missing_processed_count": max(0, planned_item_count - len(detail_rows)),
    }
    run_progress["duplicate_child_rows"] = max(0, len(detail_rows) - run_progress["detail_unique_child_count"])
    run_progress["quantity_check_ok"] = completed and len(detail_rows) == planned_item_count and len(snapshot) == run_progress["detail_unique_child_count"]
    persist_run_progress(paths, run_progress)
    write_json(paths.cache_dir / f"run_progress_{DATE}_{STAMP}.json", run_progress)
    write_self_check_report(paths, detail_rows, exception_rows, run_progress)
    write_json(paths.cache_dir / f"raw_results_{DATE}_{STAMP}.json", snapshot)
    snapshot_quality = "low_quality"
    if is_filtered_run:
        write_json(paths.cache_dir / f"snapshot_filtered_{DATE}_{STAMP}.json", snapshot)
        snapshot_quality = "filtered"
    elif project_success and len(items) == full_item_count and should_update_snapshot(snapshot, len(items)):
        write_json(paths.snapshot_file, snapshot)
        write_json(paths.cache_dir / f"snapshot_{DATE}_{STAMP}.json", snapshot)
        snapshot_quality = "snapshot"
    else:
        write_json(paths.cache_dir / f"snapshot_low_quality_{DATE}_{STAMP}.json", snapshot)
    return {
        "project": paths.name,
        "input_file": str(paths.input_file),
        "script_path": str(Path(__file__)),
        "run_mode": run_mode_label(is_filtered_run),
        "full_item_count": full_item_count,
        "planned_item_count": planned_item_count,
        "planned_unique_child_count": planned_unique_child_count,
        "processed_count": len(detail_rows),
        "missing_processed_count": run_progress["missing_processed_count"],
        "ok_count": run_progress["ok_count"],
        "error_count": run_progress["error_count"],
        "capture_failure_count": run_progress["capture_failure_count"],
        "link_invalid_count": run_progress["link_invalid_count"],
        "captcha_count": run_progress["captcha_count"],
        "blocked_count": run_progress["blocked_count"],
        "abnormal_item_count": run_progress["abnormal_item_count"],
        "exception_issue_count": run_progress["exception_issue_count"],
        "zipcode_dependent_buybox": zipcode_dependent_buybox,
        "quantity_check_ok": run_progress["quantity_check_ok"],
        "stopped_early": run_progress["stopped_early"],
        "filtered_run": is_filtered_run,
        "detail_workbook": str(detail_book_path),
        "exception_workbook": str(exception_book_path),
        "snapshot_quality": snapshot_quality,
        "snapshot_entries": len(snapshot),
        "batch_size": ASIN_BATCH_SIZE,
        "batch_count": len(batches),
        "batch_worker_count": worker_count,
        "batch_progress_file": str(latest_aggregate_file),
        "batch_run_progress_file": str(aggregate_file),
        "average_capture_seconds": aggregate["average_capture_seconds"],
        "max_capture_seconds": aggregate["max_capture_seconds"],
        "resumed_batch_count": sum(bool(state.get("resumed")) for state in aggregate["batches"].values()),
        "failed_batch_count": aggregate["failed_batch_count"],
        "project_success": project_success,
    }


def main() -> int:
    projects = discover_projects()
    if not projects:
        desktops = " or ".join(str(path) for path in DESKTOP_CANDIDATES)
        console_log("MAIN", f"No project inputs found. Put *-ASIN检查基础信息.xlsx on {desktops} to create a project.", "ERROR")
        return 0
    if not SKIP_PREFLIGHT:
        sample_url = first_project_url(projects[0])
        console_log("PREFLIGHT", f"Amazon sample={sample_url}")
        try:
            amazon_access_precheck(sample_url)
        except Exception as exc:
            error_code, action_needed = classify_precheck_failure(exc)
            write_root_summary(
                {
                    "checked_at": datetime.now().isoformat(timespec="seconds"),
                    "date": DATE,
                    "script_path": str(Path(__file__)),
                    "project_count": len(projects),
                    "projects": [],
                    "success": False,
                    "failure_stage": "precheck",
                    "error_code": error_code,
                    "sample_url": sample_url,
                    "error": str(exc),
                    "action_needed": action_needed,
                }
            )
            raise
    summaries = []
    for paths in projects:
        console_log("MAIN", f"running project={paths.name}")
        try:
            summary = run_project(paths)
        except Exception as exc:
            write_root_summary(
                {
                    "checked_at": datetime.now().isoformat(timespec="seconds"),
                    "date": DATE,
                    "script_path": str(Path(__file__)),
                    "project_count": len(projects),
                    "completed_project_count": len(summaries),
                    "projects": summaries,
                    "success": False,
                    "failure_stage": "project_run",
                    "failed_project": paths.name,
                    "error": str(exc),
                    "action_needed": "Resolve the failed project before running any later project.",
                }
            )
            raise
        summaries.append(summary)
        if not summary.get("project_success"):
            error = (
                f"[{paths.name}] project did not pass: processed={summary.get('processed_count')}/"
                f"{summary.get('planned_item_count')}, errors={summary.get('error_count')}, "
                f"capture_failures={summary.get('capture_failure_count')}, invalid_links={summary.get('link_invalid_count')}, "
                f"captcha={summary.get('captcha_count')}, blocked={summary.get('blocked_count')}, "
                f"failed_batches={summary.get('failed_batch_count')}"
            )
            write_root_summary(
                {
                    "checked_at": datetime.now().isoformat(timespec="seconds"),
                    "date": DATE,
                    "script_path": str(Path(__file__)),
                    "project_count": len(projects),
                    "completed_project_count": len(summaries) - 1,
                    "projects": summaries,
                    "success": False,
                    "failure_stage": "project_quality_gate",
                    "failed_project": paths.name,
                    "error": error,
                    "action_needed": "Resolve the failed project before running any later project.",
                }
            )
            raise RuntimeError(error)
        console_log("MAIN", f"project completed={paths.name} output={paths.output_dir}")
    write_root_summary(
        {
            "checked_at": datetime.now().isoformat(timespec="seconds"),
            "date": DATE,
            "script_path": str(Path(__file__)),
            "project_count": len(summaries),
            "projects": summaries,
            "success": True,
        },
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
